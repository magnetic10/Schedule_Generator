from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side

from app.schemas import (
    RepairEdit,
    ScheduleRepairRequest,
    ScheduleRequest,
    ScheduleSettings,
    ScheduleResult,
    WorkerInput,
    WorkerScheduleRow,
)
from app.service import (
    build_leave_need_guide,
    build_month_info,
    build_sidebar_defaults,
    build_workers,
    default_excel_filename,
    display_worker_name,
    export_schedule_result_to_excel,
    load_template_workers,
    repair_request,
    reset_worker_schedules,
    solve_request,
)
from core.config import SHIFT_DAY, SHIFT_LEAVE, SHIFT_NIGHT, SHIFT_OFF, SHIFT_OFF_NIGHT
from core.data_io import (
    _adjust_merged_ranges_for_col_delete,
    _adjust_merged_ranges_for_col_insert,
    _adjust_merged_ranges_for_row_insert,
    _finalize_basic_template_layout,
    _middle_date_col,
    _middle_worker_delete_row,
    _middle_worker_insert_row,
    _normalize_worker_footer_separator,
)
from core.models import Worker
from core.scheduler import (
    DAY_WORKLOAD_GAP_PENALTY,
    PREFERENCE_MISMATCH_PENALTY,
    SchedulerError,
    SOLVER_MAX_TIME_SECONDS,
    TO_TARGET_DEVIATION_PENALTY,
    solve_schedule,
)


class ScheduleServiceTests(unittest.TestCase):
    def _zero_night_to_cfg(self, max_day: int = 1) -> dict:
        return {
            "tgt_d": 0,
            "tgt_n": 0,
            "allow_min_d": 0,
            "allow_max_d": max_day,
            "allow_min_n": 0,
            "allow_max_n": 0,
        }

    def _max_consecutive(self, row: list, value: object) -> int:
        longest = 0
        current = 0
        for item in row:
            if item == value:
                current += 1
                longest = max(longest, current)
            else:
                current = 0
        return longest

    def test_month_info_uses_korean_holidays_for_default_target_hours(self) -> None:
        info = build_month_info(2026, 1)

        self.assertEqual(info.days_in_month, 31)
        self.assertEqual(info.workday_count, 21)
        self.assertEqual(info.default_target_hours, 168)
        self.assertEqual(info.label, "2026년 1월 | 총 31일 | 평일 21일")

    def test_sidebar_defaults_reset_settings_but_not_workers(self) -> None:
        defaults = build_sidebar_defaults(2026, 6)

        self.assertEqual(defaults.common_target_hours, 168)
        self.assertEqual(defaults.settings.target_day, 1)
        self.assertEqual(defaults.settings.target_night, 2)
        self.assertFalse(defaults.use_individual_targets)
        self.assertFalse(defaults.use_day_only)
        self.assertFalse(defaults.use_work_period)
        self.assertTrue(defaults.show_shift_colors)

    def test_reset_worker_schedules_clears_only_date_shifts(self) -> None:
        workers = [
            WorkerInput(
                name="홍길동",
                start_day=2,
                end_day=10,
                is_day_only=True,
                target_hours=120,
                preference="day",
                fixed_shifts={1: "day", 2: "leave"},
            )
        ]

        reset_workers = reset_worker_schedules(workers)

        self.assertEqual(reset_workers[0].name, "홍길동")
        self.assertEqual(reset_workers[0].start_day, 2)
        self.assertEqual(reset_workers[0].end_day, 10)
        self.assertTrue(reset_workers[0].is_day_only)
        self.assertEqual(reset_workers[0].target_hours, 120)
        self.assertEqual(reset_workers[0].preference, "day")
        self.assertEqual(reset_workers[0].fixed_shifts, {})
        self.assertEqual(workers[0].fixed_shifts, {1: "day", 2: "leave"})

    def test_build_workers_maps_stable_input_to_core_worker(self) -> None:
        workers = build_workers(
            [
                WorkerInput(
                    name="홍길동",
                    start_day=2,
                    end_day=5,
                    is_day_only=True,
                    target_hours=120,
                    preference="day",
                    fixed_shifts={2: "day", 3: "leave", 4: "교육"},
                )
            ],
            days_in_month=30,
        )

        worker = workers[0]
        self.assertEqual(worker.name, "홍길동")
        self.assertEqual(worker.start_day, 1)
        self.assertEqual(worker.end_day, 4)
        self.assertTrue(worker.is_only_day)
        self.assertEqual(worker.target_hours, 120)
        self.assertEqual(worker.shift_preference, "day")
        self.assertEqual(worker.fixed_day_days, [1])
        self.assertEqual(worker.fixed_leave_days, [2])
        self.assertEqual(worker.custom_shifts, {"교육": [3]})

    def test_solve_request_returns_schedule_and_counts(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name=f"근무자{i + 1}", target_hours=0) for i in range(8)],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=2,
                min_night=0,
                max_night=2,
            ),
            random_seed=0,
        )

        result = solve_request(req)

        self.assertIn(result.status, {"OPTIMAL", "FEASIBLE"})
        self.assertEqual(result.days_in_month, 30)
        self.assertEqual(len(result.rows), 8)
        self.assertEqual(len(result.rows[0].days), 30)
        self.assertEqual(len(result.day_counts), 30)
        self.assertEqual(len(result.night_counts), 30)
        self.assertTrue(all(0 <= count <= 2 for count in result.day_counts))
        self.assertTrue(all(0 <= count <= 2 for count in result.night_counts))

    def test_fixed_shift_is_preserved_in_solution(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(name="A", target_hours=8, fixed_shifts={1: "day", 2: "off"}),
                *[WorkerInput(name=f"B{i}", target_hours=0) for i in range(7)],
            ],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=2,
                min_night=0,
                max_night=2,
            ),
        )

        result = solve_request(req)

        self.assertEqual(result.rows[0].raw_days[0], SHIFT_DAY)
        self.assertEqual(result.rows[0].raw_days[1], SHIFT_OFF)

    def test_infeasible_fixed_night_night_reports_cycle_conflict(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=1,
            workers=[WorkerInput(name="A", target_hours=16, fixed_shifts={1: "night", 2: "night"})],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=0,
                min_night=0,
                max_night=1,
            ),
        )

        with self.assertRaises(SchedulerError) as ctx:
            solve_request(req)

        self.assertIn("1일 야간 다음 2일은 비번", str(ctx.exception))
        self.assertIn("야간이 지정", str(ctx.exception))

    def test_infeasible_fixed_off_night_off_night_reports_cycle_conflict(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=1,
            workers=[WorkerInput(name="A", target_hours=16, fixed_shifts={1: "off_night", 2: "off_night"})],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=0,
                min_night=0,
                max_night=0,
            ),
        )

        with self.assertRaises(SchedulerError) as ctx:
            solve_request(req)

        self.assertIn("1일 비번 다음 2일은 휴무", str(ctx.exception))
        self.assertIn("비번이 지정", str(ctx.exception))

    def test_solver_does_not_extend_user_fixed_off_run(self) -> None:
        worker = Worker(
            id=0,
            name="A",
            target_hours=24,
            fixed_off_days=[1, 2],
        )

        schedule, _, _ = solve_schedule(
            [worker],
            6,
            self._zero_night_to_cfg(),
            random_seed=1,
        )

        self.assertEqual(schedule[0][1], SHIFT_OFF)
        self.assertEqual(schedule[0][2], SHIFT_OFF)
        self.assertNotEqual(schedule[0][0], SHIFT_OFF)
        self.assertNotEqual(schedule[0][3], SHIFT_OFF)

    def test_objective_prioritizes_to_target_over_soft_preferences(self) -> None:
        self.assertGreater(TO_TARGET_DEVIATION_PENALTY, DAY_WORKLOAD_GAP_PENALTY)
        self.assertGreater(TO_TARGET_DEVIATION_PENALTY, PREFERENCE_MISMATCH_PENALTY)
        self.assertGreaterEqual(SOLVER_MAX_TIME_SECONDS, 40.0)

    def test_repair_request_preserves_user_edit_and_minimizes_changes(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", target_hours=16)],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=0,
            ),
        )
        baseline = ScheduleResult(
            year=2026,
            month=6,
            days_in_month=30,
            status="OPTIMAL",
            rows=[
                WorkerScheduleRow(
                    name="A",
                    days=["주", "휴", "주"] + ["휴"] * 27,
                    raw_days=[SHIFT_DAY, SHIFT_OFF, SHIFT_DAY] + [SHIFT_OFF] * 27,
                )
            ],
            day_counts=[1, 0, 1] + [0] * 27,
            night_counts=[0] * 30,
        )

        result = repair_request(
            ScheduleRepairRequest(
                request=req,
                result=baseline,
                edits=[RepairEdit(worker_index=0, day=1, shift="off")],
            )
        )

        self.assertEqual(result.rows[0].raw_days[0], SHIFT_OFF)
        self.assertEqual(result.rows[0].raw_days.count(SHIFT_DAY), 2)
        self.assertEqual(result.repair_changed_count, 2)
        self.assertTrue(
            any(
                cell["worker_index"] == 0 and cell["day"] == 1 and cell["user_locked"]
                for cell in result.repair_changed_cells
            )
        )

    def test_repair_request_rejects_edit_conflicting_with_input_fixed_shift(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", target_hours=8, fixed_shifts={1: "day"})],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=0,
            ),
        )
        baseline = ScheduleResult(
            year=2026,
            month=6,
            days_in_month=30,
            status="OPTIMAL",
            rows=[
                WorkerScheduleRow(
                    name="A",
                    days=["주"] + ["휴"] * 29,
                    raw_days=[SHIFT_DAY] + [SHIFT_OFF] * 29,
                )
            ],
            day_counts=[1] + [0] * 29,
            night_counts=[0] * 30,
        )

        with self.assertRaises(SchedulerError) as ctx:
            repair_request(
                ScheduleRepairRequest(
                    request=req,
                    result=baseline,
                    edits=[RepairEdit(worker_index=0, day=1, shift="off")],
                )
            )

        self.assertIn("입력표에서 고정한 근무", str(ctx.exception))

    def test_repair_request_can_edit_non_day_only_worker_to_night(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", target_hours=16)],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=1,
            ),
        )
        baseline = ScheduleResult(
            year=2026,
            month=6,
            days_in_month=30,
            status="OPTIMAL",
            rows=[
                WorkerScheduleRow(
                    name="A",
                    days=["주", "주"] + ["휴"] * 28,
                    raw_days=[SHIFT_DAY, SHIFT_DAY] + [SHIFT_OFF] * 28,
                )
            ],
            day_counts=[1, 1] + [0] * 28,
            night_counts=[0] * 30,
        )

        result = repair_request(
            ScheduleRepairRequest(
                request=req,
                result=baseline,
                edits=[RepairEdit(worker_index=0, day=1, shift="night")],
            )
        )

        self.assertEqual(result.rows[0].raw_days[0], SHIFT_NIGHT)
        self.assertEqual(result.rows[0].raw_days[1], SHIFT_OFF_NIGHT)
        self.assertEqual(result.repair_changed_count, 2)

    def test_repair_request_rejects_day_only_worker_to_night(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", target_hours=16, is_day_only=True)],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=1,
            ),
        )
        baseline = ScheduleResult(
            year=2026,
            month=6,
            days_in_month=30,
            status="OPTIMAL",
            rows=[
                WorkerScheduleRow(
                    name="A",
                    days=["주", "주"] + ["휴"] * 28,
                    raw_days=[SHIFT_DAY, SHIFT_DAY] + [SHIFT_OFF] * 28,
                )
            ],
            day_counts=[1, 1] + [0] * 28,
            night_counts=[0] * 30,
        )

        with self.assertRaises(SchedulerError) as ctx:
            repair_request(
                ScheduleRepairRequest(
                    request=req,
                    result=baseline,
                    edits=[RepairEdit(worker_index=0, day=1, shift="night")],
                )
            )

        self.assertIn("주간 전담 근무자는", str(ctx.exception))
        self.assertIn("야간/비번", str(ctx.exception))

    def test_night_dedicated_worker_cannot_receive_day(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", target_hours=16, dedicated_shift="night")],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=1,
            ),
        )

        result = solve_request(req)

        self.assertNotIn(SHIFT_DAY, result.rows[0].raw_days)

    def test_repair_request_rejects_night_dedicated_worker_to_day(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", target_hours=16, dedicated_shift="night")],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=1,
            ),
        )
        baseline = ScheduleResult(
            year=2026,
            month=6,
            days_in_month=30,
            status="OPTIMAL",
            rows=[
                WorkerScheduleRow(
                    name="A",
                    days=["야", "비"] + ["휴"] * 28,
                    raw_days=[SHIFT_NIGHT, SHIFT_OFF_NIGHT] + [SHIFT_OFF] * 28,
                )
            ],
            day_counts=[0] * 30,
            night_counts=[1, 0] + [0] * 28,
        )

        with self.assertRaises(SchedulerError) as ctx:
            repair_request(
                ScheduleRepairRequest(
                    request=req,
                    result=baseline,
                    edits=[RepairEdit(worker_index=0, day=1, shift="day")],
                )
            )

        self.assertIn("야간 전담 근무자는", str(ctx.exception))
        self.assertIn("주간", str(ctx.exception))

    def test_solver_allows_user_fixed_five_day_off_run(self) -> None:
        worker = Worker(
            id=0,
            name="A",
            target_hours=0,
            fixed_off_days=[0, 1, 2, 3, 4],
        )

        schedule, _, info = solve_schedule(
            [worker],
            5,
            self._zero_night_to_cfg(max_day=0),
        )

        self.assertEqual(schedule[0], [SHIFT_OFF] * 5)
        self.assertFalse(info["long_off_streak_fallback_used"])

    def test_solver_uses_five_day_auto_off_streak_only_as_fallback(self) -> None:
        worker = Worker(id=0, name="A", target_hours=0)

        schedule, _, info = solve_schedule(
            [worker],
            5,
            self._zero_night_to_cfg(max_day=0),
        )

        self.assertEqual(schedule[0], [SHIFT_OFF] * 5)
        self.assertTrue(info["long_off_streak_fallback_used"])

    def test_solver_counts_fixed_leave_inside_rest_streak(self) -> None:
        worker = Worker(
            id=0,
            name="A",
            target_hours=8,
            fixed_leave_days=[2],
        )

        schedule, _, info = solve_schedule(
            [worker],
            5,
            self._zero_night_to_cfg(max_day=0),
        )

        self.assertEqual(
            schedule[0],
            [SHIFT_OFF, SHIFT_OFF, SHIFT_LEAVE, SHIFT_OFF, SHIFT_OFF],
        )
        self.assertTrue(info["long_off_streak_fallback_used"])

    def test_solver_allows_double_night_cycle_only_when_option_is_enabled(self) -> None:
        worker = Worker(
            id=0,
            name="A",
            target_hours=32,
            fixed_night_days=[0, 2],
            fixed_off_night_days=[1, 3],
            fixed_off_days=[4, 5],
        )
        cfg = {
            "tgt_d": 0,
            "tgt_n": 0,
            "allow_min_d": 0,
            "allow_max_d": 0,
            "allow_min_n": 0,
            "allow_max_n": 1,
        }

        with self.assertRaises(SchedulerError):
            solve_schedule([worker], 6, cfg)

        schedule, _, info = solve_schedule(
            [worker],
            6,
            {**cfg, "allow_double_night_cycle": True},
        )

        self.assertEqual(
            schedule[0],
            [SHIFT_NIGHT, SHIFT_OFF_NIGHT, SHIFT_NIGHT, SHIFT_OFF_NIGHT, SHIFT_OFF, SHIFT_OFF],
        )
        self.assertTrue(info["double_night_cycle_used"])

    def test_solver_does_not_allow_double_night_cycle_from_month_start_off_night(self) -> None:
        worker = Worker(
            id=0,
            name="A",
            target_hours=24,
            fixed_off_night_days=[0, 2],
            fixed_night_days=[1],
            fixed_off_days=[3, 4],
        )

        with self.assertRaises(SchedulerError):
            solve_schedule(
                [worker],
                5,
                {
                    "tgt_d": 0,
                    "tgt_n": 0,
                    "allow_min_d": 0,
                    "allow_max_d": 0,
                    "allow_min_n": 0,
                    "allow_max_n": 1,
                    "allow_double_night_cycle": True,
                },
            )

    def test_solver_allows_fixed_leave_after_double_night_cycle_when_enabled(self) -> None:
        worker = Worker(
            id=0,
            name="A",
            target_hours=48,
            fixed_night_days=[0, 2],
            fixed_off_night_days=[1, 3],
            fixed_leave_days=[4, 5],
        )

        schedule, _, info = solve_schedule(
            [worker],
            6,
            {
                "tgt_d": 0,
                "tgt_n": 0,
                "allow_min_d": 0,
                "allow_max_d": 0,
                "allow_min_n": 0,
                "allow_max_n": 1,
                "allow_double_night_cycle": True,
                "allow_leave_after_off_night": True,
            },
        )

        self.assertEqual(
            schedule[0],
            [SHIFT_NIGHT, SHIFT_OFF_NIGHT, SHIFT_NIGHT, SHIFT_OFF_NIGHT, SHIFT_LEAVE, SHIFT_LEAVE],
        )
        self.assertTrue(info["double_night_cycle_used"])

    def test_solver_does_not_extend_user_fixed_leave_off_run(self) -> None:
        worker = Worker(
            id=0,
            name="A",
            target_hours=32,
            fixed_leave_days=[1],
            fixed_off_days=[2],
        )

        schedule, _, info = solve_schedule(
            [worker],
            6,
            self._zero_night_to_cfg(),
            random_seed=3,
        )

        self.assertEqual(schedule[0][1], SHIFT_LEAVE)
        self.assertEqual(schedule[0][2], SHIFT_OFF)
        self.assertNotEqual(schedule[0][0], SHIFT_OFF)
        self.assertNotEqual(schedule[0][3], SHIFT_OFF)
        self.assertFalse(info["long_off_streak_fallback_used"])

    def test_solve_request_reports_long_off_streaks_when_fallback_used(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=2,
            workers=[WorkerInput(name="A", target_hours=0)],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=0,
                min_night=0,
                max_night=0,
            ),
        )

        result = solve_request(req)

        self.assertTrue(result.long_off_streak_fallback_used)
        self.assertEqual(len(result.long_off_streaks), 1)
        self.assertEqual(result.long_off_streaks[0]["worker_name"], "A")
        self.assertEqual(result.long_off_streaks[0]["start_day"], 1)
        self.assertEqual(result.long_off_streaks[0]["end_day"], 28)
        self.assertEqual(result.long_off_streaks[0]["length"], 28)
        self.assertEqual(result.long_off_streaks[0]["fixed_days"], [])
        self.assertEqual(result.long_off_streaks[0]["auto_days"], list(range(1, 29)))

    def test_solve_request_reports_fixed_leave_inside_long_rest_streak(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=2,
            workers=[WorkerInput(name="A", target_hours=8, fixed_shifts={3: "leave"})],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=0,
                min_night=0,
                max_night=0,
            ),
        )

        result = solve_request(req)

        self.assertTrue(result.long_off_streak_fallback_used)
        self.assertEqual(len(result.long_off_streaks), 1)
        self.assertEqual(result.long_off_streaks[0]["fixed_days"], [3])
        self.assertEqual(result.long_off_streaks[0]["fixed_leave_days"], [3])
        self.assertNotIn(3, result.long_off_streaks[0]["auto_days"])

    def test_solver_avoids_three_day_auto_off_streak_when_possible(self) -> None:
        worker = Worker(id=0, name="A", target_hours=24)

        schedule, _, info = solve_schedule(
            [worker],
            6,
            self._zero_night_to_cfg(),
            random_seed=2,
        )

        self.assertLessEqual(self._max_consecutive(schedule[0], SHIFT_OFF), 2)
        self.assertFalse(info["long_off_streak_fallback_used"])

    def test_blank_worker_name_gets_display_label_only(self) -> None:
        self.assertEqual(display_worker_name("", 0), "A")
        self.assertEqual(display_worker_name("   ", 1), "B")
        self.assertEqual(display_worker_name("", 26), "AA")

        workers = build_workers([WorkerInput(name="", target_hours=0)], days_in_month=30)
        self.assertEqual(workers[0].name, "")

    def test_solver_does_not_auto_assign_leave(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", target_hours=8)],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=1,
            ),
        )

        result = solve_request(req)

        self.assertNotIn(SHIFT_LEAVE, result.rows[0].raw_days)

    def test_target_hours_are_hard_constraint(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(
                    name="A",
                    start_day=1,
                    end_day=1,
                    target_hours=8,
                    fixed_shifts={1: "off"},
                )
            ],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=1,
            ),
        )

        with self.assertRaises(SchedulerError):
            solve_request(req)

    def test_target_hours_cannot_exceed_work_period_capacity(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", start_day=1, end_day=18, target_hours=160)],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=0,
            ),
        )

        with self.assertRaises(SchedulerError) as ctx:
            solve_request(req)

        message = str(ctx.exception)
        self.assertIn("근무 가능 기간은 1~18일, 총 18일", message)
        self.assertIn("목표시간 160시간", message)
        self.assertIn("근무 시간이 초과", message)

    def test_leave_guide_reports_target_hours_over_work_period_capacity(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", start_day=1, end_day=18, target_hours=160)],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=0,
            ),
        )

        with self.assertRaises(SchedulerError) as ctx:
            build_leave_need_guide(req)

        self.assertIn("근무 시간이 초과", str(ctx.exception))

    def test_night_and_off_night_count_as_eight_hours_each(self) -> None:
        night_req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", target_hours=8, fixed_shifts={30: "night"})],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=1,
            ),
        )
        night_result = solve_request(night_req)
        self.assertEqual(night_result.rows[0].raw_days[29], SHIFT_NIGHT)

        off_night_req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(name="A", target_hours=8, fixed_shifts={1: "off_night"}),
            ],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=1,
            ),
        )
        off_night_result = solve_request(off_night_req)
        self.assertEqual(off_night_result.rows[0].raw_days[0], SHIFT_OFF_NIGHT)

    def test_work_period_outside_days_are_blank_in_result(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(
                    name="A",
                    start_day=2,
                    end_day=2,
                    target_hours=8,
                    fixed_shifts={2: "day"},
                )
            ],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=1,
            ),
        )

        result = solve_request(req)

        self.assertEqual(result.rows[0].raw_days[0], "")
        self.assertEqual(result.rows[0].raw_days[1], SHIFT_DAY)
        self.assertEqual(result.rows[0].raw_days[2], "")

    def test_off_night_on_first_day_allows_previous_month_cycle(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[WorkerInput(name="A", target_hours=8, fixed_shifts={1: "off_night"})],
            settings=ScheduleSettings(
                target_day=0,
                target_night=0,
                min_day=0,
                max_day=1,
                min_night=0,
                max_night=1,
            ),
        )

        result = solve_request(req)

        self.assertEqual(result.rows[0].raw_days[0], SHIFT_OFF_NIGHT)
        self.assertEqual(result.rows[0].raw_days[1], SHIFT_OFF)

    def test_leave_after_off_night_requires_option(self) -> None:
        base_settings = dict(
            target_day=0,
            target_night=0,
            min_day=0,
            max_day=1,
            min_night=0,
            max_night=1,
        )
        req_without_option = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(
                    name="A",
                    target_hours=16,
                    fixed_shifts={1: "off_night", 2: "leave"},
                )
            ],
            settings=ScheduleSettings(**base_settings),
        )

        with self.assertRaises(SchedulerError):
            solve_request(req_without_option)

        req_with_option = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(
                    name="A",
                    target_hours=16,
                    fixed_shifts={1: "off_night", 2: "leave"},
                )
            ],
            settings=ScheduleSettings(
                **base_settings,
                allow_leave_after_off_night=True,
            ),
        )

        result = solve_request(req_with_option)
        self.assertEqual(result.rows[0].raw_days[0], SHIFT_OFF_NIGHT)
        self.assertEqual(result.rows[0].raw_days[1], SHIFT_LEAVE)

    def test_emergency_range_is_not_used_when_base_range_is_enough(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(name="A", target_hours=8, fixed_shifts={1: "day"}),
                WorkerInput(name="B", target_hours=8, fixed_shifts={1: "day"}),
                WorkerInput(name="C", target_hours=0),
            ],
            settings=ScheduleSettings(
                target_day=2,
                target_night=0,
                min_day=0,
                max_day=2,
                min_night=0,
                max_night=0,
                use_emergency_range=True,
                emergency_min_day=0,
                emergency_max_day=3,
                emergency_min_night=0,
                emergency_max_night=0,
            ),
        )

        result = solve_request(req)

        self.assertEqual(result.day_counts[0], 2)
        self.assertEqual(result.emergency_days, [])

    def test_emergency_range_is_used_only_on_required_days(self) -> None:
        req_without_emergency = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(name="A", target_hours=8, fixed_shifts={1: "day"}),
                WorkerInput(name="B", target_hours=8, fixed_shifts={1: "day"}),
                WorkerInput(name="C", target_hours=8, fixed_shifts={1: "day"}),
            ],
            settings=ScheduleSettings(
                target_day=2,
                target_night=0,
                min_day=0,
                max_day=2,
                min_night=0,
                max_night=0,
            ),
        )

        with self.assertRaises(SchedulerError):
            solve_request(req_without_emergency)

        req_with_emergency = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(name="A", target_hours=8, fixed_shifts={1: "day"}),
                WorkerInput(name="B", target_hours=8, fixed_shifts={1: "day"}),
                WorkerInput(name="C", target_hours=8, fixed_shifts={1: "day"}),
            ],
            settings=ScheduleSettings(
                target_day=2,
                target_night=0,
                min_day=0,
                max_day=2,
                min_night=0,
                max_night=0,
                use_emergency_range=True,
                emergency_min_day=0,
                emergency_max_day=3,
                emergency_min_night=0,
                emergency_max_night=0,
            ),
        )

        result = solve_request(req_with_emergency)

        self.assertEqual(result.day_counts[0], 3)
        self.assertEqual(result.emergency_days, [1])

    def test_leave_need_guide_uses_individual_target_sum(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(name="A", target_hours=200),
                WorkerInput(name="B", target_hours=160),
            ],
            settings=ScheduleSettings(
                max_day=1,
                max_night=0,
            ),
        )

        guide = build_leave_need_guide(req)

        self.assertTrue(guide.uses_individual_targets)
        self.assertEqual(guide.total_target_hours, 360)
        self.assertEqual(guide.max_regular_capacity_hours, 240)
        self.assertEqual(guide.shortage_hours, 120)
        self.assertEqual(guide.suggested_leave_days, 15)

    def test_leave_need_guide_subtracts_leave_and_special_credits(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(name="A", target_hours=200, fixed_shifts={1: "leave"}),
                WorkerInput(name="B", target_hours=160, fixed_shifts={1: "교육"}),
            ],
            settings=ScheduleSettings(
                max_day=1,
                max_night=0,
                special_shifts={"교육": 16},
            ),
        )

        guide = build_leave_need_guide(req)

        self.assertEqual(guide.leave_credit_hours, 8)
        self.assertEqual(guide.special_credit_hours, 16)
        self.assertEqual(guide.non_to_credit_hours, 24)
        self.assertEqual(guide.remaining_regular_target_hours, 336)
        self.assertEqual(guide.shortage_hours, 96)
        self.assertEqual(guide.suggested_leave_days, 12)

    def test_leave_need_guide_counts_night_and_generated_off_night_capacity(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(name="A", target_hours=240),
                WorkerInput(name="B", target_hours=232),
            ],
            settings=ScheduleSettings(
                max_day=0,
                max_night=1,
            ),
        )

        guide = build_leave_need_guide(req)

        self.assertEqual(guide.max_regular_capacity_hours, 472)
        self.assertEqual(guide.shortage_hours, 0)
        self.assertEqual(guide.suggested_leave_days, 0)

    def test_leave_need_guide_uses_emergency_max_capacity_when_enabled(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(name="A", target_hours=200),
                WorkerInput(name="B", target_hours=160),
            ],
            settings=ScheduleSettings(
                max_day=1,
                max_night=0,
                use_emergency_range=True,
                emergency_max_day=2,
                emergency_max_night=0,
            ),
        )

        guide = build_leave_need_guide(req)

        self.assertTrue(guide.uses_emergency_range)
        self.assertEqual(guide.capacity_basis_label, "예외 범위 포함")
        self.assertEqual(guide.max_regular_capacity_hours, 480)
        self.assertEqual(guide.shortage_hours, 0)
        self.assertEqual(guide.suggested_leave_days, 0)
        self.assertIn("예외 범위 포함", guide.message)

    def test_leave_need_guide_rejects_unknown_special_shift_credit(self) -> None:
        req = ScheduleRequest(
            year=2026,
            month=6,
            workers=[
                WorkerInput(name="A", target_hours=160, fixed_shifts={1: "교육"}),
            ],
            settings=ScheduleSettings(
                max_day=1,
                max_night=0,
                special_shifts={},
            ),
        )

        with self.assertRaises(SchedulerError):
            build_leave_need_guide(req)

    def test_load_template_workers_without_upload_uses_default_blank_rows(self) -> None:
        result = load_template_workers(None)

        self.assertEqual(len(result.workers), 8)
        self.assertEqual(result.recognized_count, 0)
        self.assertIsNone(result.warning)
        self.assertTrue(all(worker.name == "" for worker in result.workers))

    def test_load_template_workers_uses_recognized_count_without_padding(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            template_path = Path(tmp_dir) / "template.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "2026년 6월"
            for idx in range(1, 7):
                ws.cell(row=2, column=idx + 2, value=idx)
            for idx, day in enumerate(["월", "화", "수", "목", "금", "토"], start=3):
                ws.cell(row=3, column=idx, value=day)
            ws["B4"] = "홍길동"
            ws["C4"] = "주"
            ws["B5"] = "김철수"
            ws["C5"] = "야"
            wb.save(template_path)
            wb.close()

            result = load_template_workers(str(template_path))

        self.assertEqual(result.recognized_count, 2)
        self.assertEqual([worker.name for worker in result.workers], ["홍길동", "김철수"])
        self.assertEqual(len(result.workers), 2)
        self.assertIsNone(result.warning)
        self.assertEqual(result.template_info["worker_row_start"], 4)

    def test_load_template_workers_fallback_does_not_pad_recognized_names(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            template_path = Path(tmp_dir) / "fallback.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["B4"] = "홍길동"
            ws["B5"] = "김철수"
            wb.save(template_path)
            wb.close()

            result = load_template_workers(str(template_path))

        self.assertEqual(result.recognized_count, 2)
        self.assertEqual([worker.name for worker in result.workers], ["홍길동", "김철수"])
        self.assertEqual(len(result.workers), 2)
        self.assertIsNotNone(result.warning)

    def test_load_template_workers_accepts_valid_template_without_real_names(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            template_path = Path(tmp_dir) / "blank_template.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "2026년 6월 근무계획표"
            ws["A3"] = "월일"
            ws["B3"] = "6월"
            for idx in range(1, 31):
                ws.cell(row=3, column=idx + 2, value=idx)
            for idx, day in enumerate(["월", "화", "수", "목", "금", "토", "일"], start=3):
                ws.cell(row=4, column=idx, value=day)
            ws["B4"] = "요일"
            for row in range(5, 13):
                ws.cell(row=row, column=2, value=f"근무자{row - 4}")
                ws.cell(row=row, column=3, value="휴")
            ws["A13"] = "주간계"
            wb.save(template_path)
            wb.close()

            result = load_template_workers(str(template_path))

        self.assertEqual(result.recognized_count, 0)
        self.assertEqual(result.workers, [])
        self.assertTrue(result.preserve_existing_workers)
        self.assertEqual(result.template_status, "valid_without_names")
        self.assertEqual(result.template_info["date_row"], 3)
        self.assertEqual(result.template_info["worker_row_start"], 5)

    def test_export_schedule_result_to_excel_calls_core_export_with_display_names(self) -> None:
        result = ScheduleResult(
            year=2026,
            month=6,
            days_in_month=2,
            status="OPTIMAL",
            rows=[
                WorkerScheduleRow(name="A", days=["주", "교육"], raw_days=[SHIFT_DAY, "교육"]),
                WorkerScheduleRow(name="김철수", days=["휴", "야"], raw_days=[SHIFT_OFF, SHIFT_NIGHT]),
            ],
            day_counts=[1, 0],
            night_counts=[0, 1],
        )

        with patch("app.service.export_to_excel") as export_mock:
            export_result = export_schedule_result_to_excel(
                template_path="template.xlsx",
                output_path=str(Path("out") / default_excel_filename(2026, 6)),
                result=result,
                apply_shift_colors=True,
            )

        export_mock.assert_called_once()
        kwargs = export_mock.call_args.kwargs
        self.assertEqual(kwargs["template_path"], "template.xlsx")
        self.assertEqual(kwargs["num_days"], 2)
        self.assertEqual(kwargs["year"], 2026)
        self.assertEqual(kwargs["month"], 6)
        self.assertFalse(kwargs["apply_shift_colors"])
        self.assertEqual([worker.name for worker in kwargs["workers"]], ["A", "김철수"])
        self.assertEqual(kwargs["schedule"], [[SHIFT_DAY, "교육"], [SHIFT_OFF, SHIFT_NIGHT]])
        self.assertEqual(export_result.filename, default_excel_filename(2026, 6))
        self.assertEqual(export_result.worker_count, 2)

    def test_excel_merged_ranges_adjust_for_column_insert_and_delete(self) -> None:
        ranges = [(1, 1, 2, 31), (1, 32, 3, 36), (17, 1, 18, 46)]

        inserted = _adjust_merged_ranges_for_col_insert(ranges, 32)
        self.assertIn((1, 1, 2, 31), inserted)
        self.assertIn((1, 33, 3, 37), inserted)
        self.assertIn((17, 1, 18, 47), inserted)

        deleted = _adjust_merged_ranges_for_col_delete(ranges, 30)
        deleted = _adjust_merged_ranges_for_col_delete(deleted, 30)
        self.assertIn((1, 1, 2, 29), deleted)
        self.assertIn((1, 30, 3, 34), deleted)
        self.assertIn((17, 1, 18, 44), deleted)

    def test_excel_merged_ranges_adjust_for_row_insert(self) -> None:
        ranges = [(1, 37, 16, 50), (17, 1, 18, 46)]

        adjusted = _adjust_merged_ranges_for_row_insert(ranges, 13)
        self.assertIn((1, 37, 17, 50), adjusted)
        self.assertIn((18, 1, 19, 46), adjusted)

    def test_excel_worker_rows_are_adjusted_inside_worker_area(self) -> None:
        self.assertEqual(_middle_worker_insert_row(5, 8, 1), 9)
        self.assertEqual(_middle_worker_delete_row(5, 8, 1), 9)
        self.assertEqual(_middle_worker_insert_row(5, 2, 1), 6)
        self.assertEqual(_middle_worker_delete_row(5, 2, 1), 5)

    def test_excel_date_columns_are_adjusted_inside_date_area(self) -> None:
        self.assertEqual(_middle_date_col(2, 30), 17)
        self.assertEqual(_middle_date_col(2, 31), 17)

    def test_excel_worker_footer_separator_is_normalized_after_row_adjustment(self) -> None:
        wb = Workbook()
        ws = wb.active
        thin = Side(style="thin")
        medium = Side(style="medium")
        for row in range(1, 8):
            for col in range(1, 4):
                ws.cell(row, col).border = Border(top=thin, bottom=thin, left=thin, right=thin)

        for col in range(1, 4):
            worker_cell = ws.cell(5, col)
            worker_cell.border = Border(
                top=worker_cell.border.top,
                bottom=medium,
                left=worker_cell.border.left,
                right=worker_cell.border.right,
            )
            footer_cell = ws.cell(6, col)
            footer_cell.border = Border(
                top=medium,
                bottom=footer_cell.border.bottom,
                left=footer_cell.border.left,
                right=footer_cell.border.right,
            )

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "separator.xlsx"
            wb.save(output_path)
            wb.close()

            _normalize_worker_footer_separator(
                output_path=str(output_path),
                worker_row_start=2,
                worker_count=4,
                worker_row_step=1,
                footer_row=6,
            )

            updated = load_workbook(output_path)
            updated_ws = updated.active
            for col in range(1, 4):
                self.assertEqual(updated_ws.cell(5, col).border.bottom.style, "thin")
                self.assertEqual(updated_ws.cell(6, col).border.top.style, "thin")
            updated.close()

    def test_basic_template_layout_rewrites_dynamic_formulas_and_borders(self) -> None:
        wb = Workbook()
        ws = wb.active
        for row in range(1, 22):
            for col in range(1, 45):
                ws.cell(row, col).value = ""

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "basic.xlsx"
            wb.save(output_path)
            wb.close()

            _finalize_basic_template_layout(
                output_path=str(output_path),
                date_row=3,
                day_row=4,
                date_col_start=2,
                name_col=1,
                worker_row_start=5,
                worker_count=10,
                worker_row_step=1,
                worker_data_row_offset=0,
                num_days=31,
            )

            updated = load_workbook(output_path, data_only=False)
            ws = updated.active
            self.assertEqual(ws["AG4"].value, "주")
            self.assertEqual(ws["AH4"].value, "야")
            self.assertEqual(ws["AI4"].value, "비")
            self.assertEqual(ws["AJ4"].value, "휴")
            self.assertEqual(ws["AK4"].value, "연")
            self.assertEqual(ws["AL4"].value, "근무시간")
            self.assertEqual(ws["AH5"].value, '=COUNTIF(B5:AF5,"야")')
            self.assertEqual(ws["AL5"].value, "=(AG5 + AK5 + AH5 + AI5)*8")
            self.assertEqual(ws["B15"].value, '=COUNTIF(B5:B14,"주")')
            self.assertEqual(ws["B16"].value, '=COUNTIF(B5:B14,"야")')
            self.assertEqual(ws["B17"].value, '=COUNTIF(B5:B14,"비")')
            self.assertEqual(ws["B18"].value, '=COUNTIF(B5:B14,"휴")')
            self.assertEqual(ws["AF5"].border.right.style, "medium")
            self.assertEqual(ws["AG5"].border.left.style, "medium")
            self.assertEqual(ws["A14"].border.bottom.style, "medium")
            self.assertEqual(ws["A15"].border.top.style, "medium")
            updated.close()


if __name__ == "__main__":
    unittest.main()
