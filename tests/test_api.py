from __future__ import annotations

import asyncio
import io
import json
import unittest

from fastapi import UploadFile
from openpyxl import Workbook

from app.main import default_state, delete_template, repair, reset_input, solve, upload_template


class ApiTests(unittest.TestCase):
    def test_default_state_returns_defaults_and_blank_workers(self) -> None:
        data = default_state(year=2026, month=1)

        self.assertEqual(data["defaults"]["month_info"]["days_in_month"], 31)
        self.assertEqual(data["defaults"]["common_target_hours"], 168)
        self.assertEqual(len(data["workers"]), 8)
        self.assertTrue(all(worker["name"] == "" for worker in data["workers"]))

    def test_reset_input_clears_fixed_shifts_only(self) -> None:
        data = reset_input(
            {
                "workers": [
                    {
                        "name": "홍길동",
                        "start_day": 2,
                        "end_day": 10,
                        "target_hours": 120,
                        "is_day_only": True,
                        "fixed_shifts": {"1": "day", "2": "leave"},
                    }
                ]
            }
        )

        worker = data["workers"][0]
        self.assertEqual(worker["name"], "홍길동")
        self.assertEqual(worker["start_day"], 2)
        self.assertEqual(worker["end_day"], 10)
        self.assertEqual(worker["target_hours"], 120)
        self.assertTrue(worker["is_day_only"])
        self.assertEqual(worker["fixed_shifts"], {})

    def test_solve_endpoint_returns_schedule_result(self) -> None:
        data = solve(
            {
                "year": 2026,
                "month": 1,
                "workers": [{"name": "", "target_hours": 0}],
                "settings": {
                    "target_day": 0,
                    "target_night": 0,
                    "min_day": 0,
                    "max_day": 0,
                    "min_night": 0,
                    "max_night": 0,
                },
            }
        )

        self.assertTrue(data["success"])
        self.assertEqual(data["result"]["rows"][0]["name"], "")
        self.assertEqual(len(data["result"]["rows"][0]["days"]), 31)

    def test_solve_endpoint_reports_target_hours_over_work_period_capacity(self) -> None:
        response = solve(
            {
                "year": 2026,
                "month": 6,
                "workers": [{"name": "A", "start_day": 1, "end_day": 18, "target_hours": 160}],
                "settings": {
                    "target_day": 0,
                    "target_night": 0,
                    "min_day": 0,
                    "max_day": 1,
                    "min_night": 0,
                    "max_night": 0,
                },
            }
        )
        body = json.loads(response.body.decode("utf-8"))

        self.assertEqual(response.status_code, 422)
        self.assertFalse(body["success"])
        self.assertIn("근무 시간이 초과", body["error"])

    def test_repair_endpoint_returns_repaired_result(self) -> None:
        data = repair(
            {
                "request": {
                    "year": 2026,
                    "month": 6,
                    "workers": [{"name": "A", "target_hours": 16}],
                    "settings": {
                        "target_day": 0,
                        "target_night": 0,
                        "min_day": 0,
                        "max_day": 1,
                        "min_night": 0,
                        "max_night": 0,
                    },
                },
                "result": {
                    "year": 2026,
                    "month": 6,
                    "days_in_month": 30,
                    "status": "OPTIMAL",
                    "rows": [
                        {
                            "name": "A",
                            "days": ["주", "휴", "주"] + ["휴"] * 27,
                            "raw_days": [0, 3, 0] + [3] * 27,
                        }
                    ],
                    "day_counts": [1, 0, 1] + [0] * 27,
                    "night_counts": [0] * 30,
                },
                "edits": [{"worker_index": 0, "day": 1, "shift": "off"}],
            }
        )

        self.assertTrue(data["success"])
        self.assertEqual(data["result"]["rows"][0]["raw_days"][0], 3)
        self.assertEqual(data["result"]["repair_changed_count"], 4)

    def test_upload_template_returns_recognized_workers_and_delete_keeps_api_simple(self) -> None:
        workbook_bytes = _template_workbook_bytes()
        upload_file = UploadFile(
            file=io.BytesIO(workbook_bytes),
            filename="template.xlsx",
        )

        data = asyncio.run(upload_template(file=upload_file))

        self.assertEqual(data["load_result"]["recognized_count"], 2)
        self.assertEqual(
            [worker["name"] for worker in data["load_result"]["workers"]],
            ["홍길동", "김철수"],
        )

        delete_result = delete_template(data["file_id"])
        self.assertTrue(delete_result["deleted"])

    def test_upload_template_rejects_wrong_xlsx_format(self) -> None:
        workbook_bytes = _wrong_workbook_bytes()
        upload_file = UploadFile(
            file=io.BytesIO(workbook_bytes),
            filename="wrong.xlsx",
        )

        response = asyncio.run(upload_template(file=upload_file))
        body = json.loads(response.body.decode("utf-8"))

        self.assertEqual(response.status_code, 422)
        self.assertFalse(body["success"])
        self.assertIn("근무표 서식으로 인식하지 못했습니다", body["error"])

    def test_upload_template_accepts_valid_template_without_real_names(self) -> None:
        workbook_bytes = _blank_template_workbook_bytes()
        upload_file = UploadFile(
            file=io.BytesIO(workbook_bytes),
            filename="blank_template.xlsx",
        )

        data = asyncio.run(upload_template(file=upload_file))

        self.assertEqual(data["load_result"]["recognized_count"], 0)
        self.assertEqual(data["load_result"]["workers"], [])
        self.assertTrue(data["load_result"]["preserve_existing_workers"])
        self.assertEqual(data["load_result"]["template_status"], "valid_without_names")
        delete_template(data["file_id"])


def _template_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "2026년 6월"
    for idx in range(1, 7):
        ws.cell(row=2, column=idx + 2, value=idx)
    for idx, day in enumerate(["월", "화", "수", "목", "금", "토"], start=3):
        ws.cell(row=3, column=idx, value=day)
    ws["B4"] = "홍길동"
    ws["C4"] = "주"
    ws["B5"] = "김철수"
    ws["C5"] = "야"

    stream = io.BytesIO()
    wb.save(stream)
    wb.close()
    stream.seek(0)
    return stream.read()


def _wrong_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "일반 문서"
    ws["A2"] = "근무표 서식이 아닌 엑셀"

    stream = io.BytesIO()
    wb.save(stream)
    wb.close()
    stream.seek(0)
    return stream.read()


def _blank_template_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "2026년 6월 근무계획표"
    ws["A3"] = "월일"
    ws["B3"] = "6월"
    ws["B4"] = "요일"
    for idx in range(1, 31):
        ws.cell(row=3, column=idx + 2, value=idx)
    for idx, day in enumerate(["월", "화", "수", "목", "금", "토", "일"], start=3):
        ws.cell(row=4, column=idx, value=day)
    for row in range(5, 13):
        ws.cell(row=row, column=2, value=f"근무자{row - 4}")
        ws.cell(row=row, column=3, value="휴")
    ws["A13"] = "주간계"

    stream = io.BytesIO()
    wb.save(stream)
    wb.close()
    stream.seek(0)
    return stream.read()


if __name__ == "__main__":
    unittest.main()
