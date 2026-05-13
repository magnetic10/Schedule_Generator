import csv
import os
import re
import datetime
import calendar
from copy import copy
from .config import SHIFT_TO_STR, NUM_WORKERS
from .models import Worker

SHIFT_KEYWORDS  = {"주", "야", "비", "휴", "연"}
KOREAN_DAYS     = {"월", "화", "수", "목", "금", "토", "일"}
FOOTER_KEYWORDS = {
    # 집계 키워드 (계 붙은 형태)
    "주간계", "야간계", "휴무계", "비번계", "주계", "야계", "휴계", "비계",
    # 합계/통계 일반 키워드
    "합계", "총원", "소계", "계",
    # 근무 유형 단독 (통계 행에 자주 쓰임)
    "주간", "야간", "휴무", "비번", "연가", "당직", "합산",
    # 기타
    "비고", "합",
}
HEADER_KEYWORDS = {
    "요일", "일자", "월일", "날짜", "성명", "이름", "근무자명", "대상자",
}
PLACEHOLDER_NAME_RE = re.compile(r"^(근무자|직원|이름|성명)\s*\d+$")
MONTH_TEXT_RE = re.compile(r"(\d{4}\s*년\s*)?\d{1,2}\s*월")
HANGUL_NAME_RE = re.compile(r"^[가-힣]{2,}(?:[·\s]?[가-힣]{1,})?$")

SHIFT_COLOR_RULES = {
    "-": {"font": "FFADB0AB", "fill": None},
    "주": {"font": "FF000000", "fill": "FF93E645"},
    "야": {"font": "FF000000", "fill": "FFE68845"},
    "비": {"font": "FF000000", "fill": "FF45E6E1"},
    "휴": {"font": "FF000000", "fill": "FF4558E6"},
    "연": {"font": "FF000000", "fill": "FF9245E6"},
}

# ───────────────────────────────────────────────
# 자체 열 문자 변환 (A=1, B=2, ... Z=26, AA=27, ...)
# ───────────────────────────────────────────────
def get_column_letter(n: int) -> str:
    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result

def col_letter_to_idx(letters: str) -> int:
    """'A' → 1, 'B' → 2, 'AA' → 27"""
    idx = 0
    for ch in letters.upper():
        idx = idx * 26 + (ord(ch) - 64)
    return idx


def _range_address(min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _read_merged_ranges(ws_openpyxl) -> list[tuple[int, int, int, int]]:
    return [
        (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
        for rng in ws_openpyxl.merged_cells.ranges
    ]


def _adjust_interval_for_insert(start: int, end: int, index: int, count: int) -> tuple[int, int]:
    if end < index:
        return start, end
    if start >= index:
        return start + count, end + count
    return start, end + count


def _adjust_interval_for_delete(
    start: int,
    end: int,
    index: int,
    count: int,
) -> tuple[int, int] | None:
    delete_end = index + count - 1
    remaining = [pos for pos in range(start, end + 1) if not index <= pos <= delete_end]
    if not remaining:
        return None
    shifted = [pos - count if pos > delete_end else pos for pos in remaining]
    return min(shifted), max(shifted)


def _adjust_merged_ranges_for_col_insert(
    ranges: list[tuple[int, int, int, int]],
    index: int,
    count: int = 1,
) -> list[tuple[int, int, int, int]]:
    adjusted = []
    for min_row, min_col, max_row, max_col in ranges:
        new_min_col, new_max_col = _adjust_interval_for_insert(min_col, max_col, index, count)
        adjusted.append((min_row, new_min_col, max_row, new_max_col))
    return adjusted


def _adjust_merged_ranges_for_row_insert(
    ranges: list[tuple[int, int, int, int]],
    index: int,
    count: int = 1,
) -> list[tuple[int, int, int, int]]:
    adjusted = []
    for min_row, min_col, max_row, max_col in ranges:
        new_min_row, new_max_row = _adjust_interval_for_insert(min_row, max_row, index, count)
        adjusted.append((new_min_row, min_col, new_max_row, max_col))
    return adjusted


def _adjust_merged_ranges_for_col_delete(
    ranges: list[tuple[int, int, int, int]],
    index: int,
    count: int = 1,
) -> list[tuple[int, int, int, int]]:
    adjusted = []
    for min_row, min_col, max_row, max_col in ranges:
        col_interval = _adjust_interval_for_delete(min_col, max_col, index, count)
        if col_interval is None:
            continue
        new_min_col, new_max_col = col_interval
        adjusted.append((min_row, new_min_col, max_row, new_max_col))
    return adjusted


def _adjust_merged_ranges_for_row_delete(
    ranges: list[tuple[int, int, int, int]],
    index: int,
    count: int = 1,
) -> list[tuple[int, int, int, int]]:
    adjusted = []
    for min_row, min_col, max_row, max_col in ranges:
        row_interval = _adjust_interval_for_delete(min_row, max_row, index, count)
        if row_interval is None:
            continue
        new_min_row, new_max_row = row_interval
        adjusted.append((new_min_row, min_col, new_max_row, max_col))
    return adjusted


def _unmerge_ranges_xw(ws_xw, ranges: list[tuple[int, int, int, int]]) -> None:
    for min_row, min_col, max_row, max_col in ranges:
        if min_row == max_row and min_col == max_col:
            continue
        ws_xw.api.Range(_range_address(min_row, min_col, max_row, max_col)).UnMerge()


def _remerge_ranges_xw(ws_xw, ranges: list[tuple[int, int, int, int]]) -> None:
    seen: set[str] = set()
    for min_row, min_col, max_row, max_col in ranges:
        if min_row == max_row and min_col == max_col:
            continue
        address = _range_address(min_row, min_col, max_row, max_col)
        if address in seen:
            continue
        seen.add(address)
        ws_xw.api.Range(address).Merge()


def _middle_worker_insert_row(worker_row_start: int, worker_count: int, row_step: int) -> int:
    if worker_count <= 1:
        return worker_row_start
    return worker_row_start + max(1, worker_count // 2) * row_step


def _middle_worker_delete_row(worker_row_start: int, worker_count: int, row_step: int) -> int:
    if worker_count <= 1:
        return worker_row_start
    if worker_count == 2:
        return worker_row_start
    return worker_row_start + (worker_count // 2) * row_step


def _middle_date_col(date_col_start: int, day_count: int) -> int:
    if day_count <= 1:
        return date_col_start
    return date_col_start + max(1, day_count // 2)


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_day_number(value, expected: int | None = None) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        number = int(value)
        if float(value) != float(number):
            return False
    else:
        text = _cell_text(value)
        if not text.isdigit():
            return False
        number = int(text)
    if expected is not None:
        return number == expected
    return 1 <= number <= 31


def _is_number_like(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    text = _cell_text(value)
    if not text:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def _is_placeholder_name(value: str) -> bool:
    return bool(PLACEHOLDER_NAME_RE.match(_cell_text(value).replace(" ", "")))


def _is_excluded_name_text(value: str) -> bool:
    text = _cell_text(value)
    if not text:
        return True
    if text in KOREAN_DAYS or text in FOOTER_KEYWORDS or text in HEADER_KEYWORDS:
        return True
    if MONTH_TEXT_RE.fullmatch(text):
        return True
    return False


def is_actual_worker_name(value: str) -> bool:
    text = _cell_text(value)
    if _is_excluded_name_text(text) or _is_placeholder_name(text):
        return False
    return bool(HANGUL_NAME_RE.match(text))


def _is_worker_name_candidate(value: str) -> bool:
    text = _cell_text(value)
    if _is_excluded_name_text(text):
        return False
    return is_actual_worker_name(text) or _is_placeholder_name(text)


def _is_shift_like(value) -> bool:
    text = _cell_text(value)
    if not text:
        return False
    if text in SHIFT_KEYWORDS or text in {"-", "연가"}:
        return True
    return len(text) == 1 and not text.isdigit() and text not in KOREAN_DAYS


def analyze_template_v3(ws_openpyxl) -> dict:
    ws = ws_openpyxl
    max_row = ws.max_row or 50
    max_col = ws.max_column or 50
    info = {
        "sheet_name": ws.title,
        "date_row": None,
        "day_row": None,
        "date_col_start": None,
        "name_col": None,
        "worker_row_start": None,
        "worker_row_step": 1,
        "worker_data_row_offset": 0,
        "worker_row_count": 0,
        "footer_row": None,
        "month_cells": [],
        "right_stat_col_start": None,
        "template_days": 31,
        "has_real_worker_names": False,
    }

    def val(row: int, col: int):
        return ws.cell(row=row, column=col).value

    best_date: tuple[int, int, int] | None = None
    for row in range(1, min(max_row, 20) + 1):
        for col in range(1, max_col + 1):
            if not _is_day_number(val(row, col), 1):
                continue
            length = 1
            while col + length <= max_col and _is_day_number(val(row, col + length), length + 1):
                length += 1
            if length >= 6 and (best_date is None or length > best_date[2]):
                best_date = (row, col, length)

    if best_date is None:
        raise ValueError("날짜 행을 찾을 수 없습니다. 서식 파일을 확인해 주세요.")

    info["date_row"], info["date_col_start"], detected_days = best_date
    info["template_days"] = min(31, max(28, detected_days))

    date_row = int(info["date_row"])
    date_col_start = int(info["date_col_start"])
    template_days = int(info["template_days"])
    last_date_col = date_col_start + template_days - 1

    for row in range(max(1, date_row - 2), min(max_row, date_row + 2) + 1):
        if row == date_row:
            continue
        day_count = 0
        for col in range(date_col_start, min(max_col, date_col_start + 13) + 1):
            if _cell_text(val(row, col)) in KOREAN_DAYS:
                day_count += 1
        if day_count >= 5:
            info["day_row"] = row
            break

    search_start = max(date_row, int(info["day_row"] or 0)) + 1

    def row_profile(row: int) -> dict[str, int]:
        profile = {"shift": 0, "blank": 0, "numeric": 0, "formula": 0, "weekday": 0, "other": 0}
        for col in range(date_col_start, min(max_col, last_date_col) + 1):
            raw = val(row, col)
            text = _cell_text(raw)
            if not text:
                profile["blank"] += 1
            elif text.startswith("="):
                profile["formula"] += 1
            elif _is_number_like(raw):
                profile["numeric"] += 1
            elif text in KOREAN_DAYS:
                profile["weekday"] += 1
            elif _is_shift_like(text):
                profile["shift"] += 1
            else:
                profile["other"] += 1
        return profile

    def looks_like_worker_schedule(row: int) -> bool:
        profile = row_profile(row)
        if profile["numeric"] or profile["formula"] or profile["weekday"] >= 3:
            return False
        return profile["shift"] > 0 or profile["blank"] >= max(5, template_days // 2)

    found_rows: list[tuple[int, int, str, bool]] = []
    empty_after_found = 0
    for row in range(search_start, min(max_row, search_start + 100) + 1):
        left_values = [_cell_text(val(row, col)) for col in range(1, date_col_start)]
        if any(text in FOOTER_KEYWORDS for text in left_values if text):
            if found_rows:
                info["footer_row"] = row
                break

        candidates = [
            (col, _cell_text(val(row, col)))
            for col in range(1, date_col_start)
            if _is_worker_name_candidate(_cell_text(val(row, col)))
        ]
        if not candidates or not looks_like_worker_schedule(row):
            if found_rows:
                empty_after_found += 1
                if empty_after_found >= 3:
                    break
            continue

        empty_after_found = 0
        col, name = candidates[-1]
        found_rows.append((row, col, name, is_actual_worker_name(name)))

    if not found_rows:
        raise ValueError("근무자 행을 찾을 수 없습니다. 서식 파일을 확인해 주세요.")

    info["worker_row_start"] = found_rows[0][0]
    info["name_col"] = found_rows[0][1]
    info["worker_row_count"] = len(found_rows)
    info["has_real_worker_names"] = any(item[3] for item in found_rows)

    if len(found_rows) >= 2:
        info["worker_row_step"] = max(1, found_rows[1][0] - found_rows[0][0])

    step = int(info["worker_row_step"])
    name_row = int(info["worker_row_start"])
    best_offset = 0
    best_shift_count = -1
    for offset in range(step):
        profile = row_profile(name_row + offset)
        if profile["shift"] > best_shift_count:
            best_shift_count = profile["shift"]
            best_offset = offset
    info["worker_data_row_offset"] = best_offset

    if info["footer_row"] is None:
        last_worker_row = found_rows[-1][0] + step
        for row in range(last_worker_row, min(max_row, last_worker_row + 30) + 1):
            for col in range(1, min(max_col, max(date_col_start, 16)) + 1):
                if _cell_text(val(row, col)) in FOOTER_KEYWORDS:
                    info["footer_row"] = row
                    break
            if info["footer_row"]:
                break

    for row in range(1, search_start):
        for col in range(1, max_col + 1):
            text = _cell_text(val(row, col))
            if text and MONTH_TEXT_RE.search(text):
                info["month_cells"].append((row, col))

    for col in range(last_date_col + 1, min(max_col, last_date_col + 10) + 1):
        for row in range(max(1, date_row - 2), min(max_row, date_row + 3) + 1):
            text = _cell_text(val(row, col))
            if text:
                info["right_stat_col_start"] = col
                return info

    return info

# ───────────────────────────────────────────────
# 지능형 서식 분석 엔진
# ───────────────────────────────────────────────
def analyze_template(ws_openpyxl) -> dict:
    """
    openpyxl worksheet 를 받아 서식 구조를 분석하고 아래 딕셔너리를 반환한다.
    {
      'date_row'      : int,          # 날짜(1,2,3…) 가 있는 행 번호
      'day_row'       : int | None,   # 요일(월,화…) 이 있는 행 번호 (없으면 None)
      'date_col_start': int,          # 1일이 시작되는 열 번호
      'name_col'      : int,          # 이름이 있는 열 번호
      'worker_row_start': int,        # 첫 근무자 행 번호
      'worker_row_step': int,         # 1명당 행 수 (1 또는 2)
      'worker_data_row_offset': int,  # 이름 행으로부터 실제 근무 기호가 있는 행 오프셋
      'footer_row'    : int,          # 하단 통계 시작 행
      'month_cells'   : list[tuple],  # "n월" 텍스트가 있는 (row, col) 목록
      'right_stat_col_start': int | None, # 우측 통계 시작 열 번호 (없으면 None)
      'template_days' : int,          # 원본 서식의 날짜 수 (28~31)
    }
    """
    ws = ws_openpyxl
    max_row = ws.max_row or 50
    max_col = ws.max_column or 50
    # 검토 범위: 최대 P열(16) 까지만
    SEARCH_MAX_COL = min(max_col, 16)

    info = {
        'date_row': None, 'day_row': None,
        'date_col_start': None, 'name_col': None,
        'worker_row_start': None, 'worker_row_step': 1,
        'worker_data_row_offset': 0, 'footer_row': None,
        'month_cells': [], 'right_stat_col_start': None,
        'template_days': 31,
    }

    def cell_val(r, c):
        v = ws.cell(row=r, column=c).value
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return v
        return str(v).strip()

    # ── 1. 날짜 행 & 날짜 시작 열 탐지 ──────────────────
    # 1~10행 사이에서 1,2,3... 이 연속으로 6개 이상 등장하는 행
    for r in range(1, 15):
        consecutive = 0
        first_col = None
        prev = 0
        for c in range(1, SEARCH_MAX_COL + 1):
            v = cell_val(r, c)
            if isinstance(v, (int, float)) and int(v) == prev + 1:
                consecutive += 1
                if consecutive == 1:
                    first_col = c - 1  # 이전 열이 시작점 후보... 아래서 조정
                    first_col = c  # 실제 '1'은 consecutive==1 일 때 c-1 이 아니라 c
            else:
                if consecutive >= 6:
                    break
                consecutive = 0
                prev = 0
            if isinstance(v, (int, float)):
                prev = int(v)
            else:
                prev = 0

        if consecutive >= 6:
            info['date_row'] = r
            # first_col 재탐색: 해당 행에서 값이 1인 열
            for c in range(1, SEARCH_MAX_COL + 1):
                v = cell_val(r, c)
                if isinstance(v, (int, float)) and int(v) == 1:
                    info['date_col_start'] = c
                    break
            # 원본 서식의 날짜 수 계산 (마지막으로 연속된 숫자)
            last_day = 0
            for c in range(1, max_col + 1):
                v = cell_val(r, c)
                if isinstance(v, (int, float)) and int(v) > last_day:
                    last_day = int(v)
                elif last_day >= 28 and last_day <= 31:
                    break
            info['template_days'] = max(last_day, 28)
            break

    if info['date_row'] is None:
        raise ValueError("날짜 행을 찾을 수 없습니다. 서식 파일을 확인해 주세요.")

    # ── 2. 요일 행 탐지 ──────────────────────────────────
    for r in range(max(1, info['date_row'] - 2), info['date_row'] + 3):
        if r == info['date_row']:
            continue
        day_cnt = 0
        for c in range(info['date_col_start'], info['date_col_start'] + 10):
            v = cell_val(r, c)
            if v in KOREAN_DAYS:
                day_cnt += 1
        if day_cnt >= 5:
            info['day_row'] = r
            break

    # ── 3. 이름 열 & 근무자 행 탐지 ─────────────────────
    # A~D열(1~4)에서 2글자 이상 한글 이름을 찾고,
    # 그 행 오른쪽 1~5칸 범위에 shift 키워드가 있는지 확인
    import re as _re
    HANGUL_RE = _re.compile(r'^[\uAC00-\uD7A3]{2,}$')

    def is_real_worker_row(r, name_col, date_col_start):
        """이름 오른쪽 date_col_start 까지 없거나, date_col_start 이후에 shift or 빈칸 있으면 OK"""
        for c in range(date_col_start, date_col_start + 10):
            v = cell_val(r, c)
            if v in SHIFT_KEYWORDS or v is None or v == "" or v == "-":
                return True
        return False

    found_name_rows = []  # [(row, col, name)]
    # 날짜 행과 요일 행을 모두 건너뛰고 그 다음 행부터 이름 탐색
    if info['day_row']:
        search_start_row = max(info['date_row'], info['day_row']) + 1
    else:
        search_start_row = info['date_row'] + 1
    date_col = info['date_col_start'] or 3

    for r in range(search_start_row, min(search_start_row + 60, max_row + 1)):
        # A~D열 순서로 확인
        for c in range(1, 5):
            v = cell_val(r, c)
            if not v or not isinstance(v, str):
                continue
            sv = str(v).strip()
            # 필터링: 2글자 미만이거나, 요일 단어이거나, 통계 키워드
            if len(sv) < 2:
                continue
            if sv in KOREAN_DAYS or sv in FOOTER_KEYWORDS:
                continue
            if not HANGUL_RE.match(sv):
                continue

            # 이름 오른쪽에 날짜 데이터(주/야/비/휴/빈칸)가 있는지 확인
            # date_col_start 기준으로 해당 행에 shift 키워드 또는 빈 셀 존재 시 근무자로 인정
            has_data_col = False
            for dc in range(date_col, date_col + 5):
                dv = cell_val(r, dc)
                if dv is None or dv == "" or str(dv).strip() in SHIFT_KEYWORDS or str(dv).strip() == "-":
                    has_data_col = True
                    break

            # ── 핵심 검증: 날짜 범위에 숫자가 있으면 통계 행 → 제외
            # 근무자 행: 날짜 칸에 주/야/비/휴/연 또는 빈칸(-)
            # 통계 행: 날짜 칸에 숫자(총원=11, 주간=2 등)
            is_stats_row = False
            has_any_shift = False
            for dc in range(date_col, date_col + 15):
                dv = cell_val(r, dc)
                if dv is None or dv == "" or str(dv).strip() in ("", "-"):
                    continue
                if str(dv).strip() in SHIFT_KEYWORDS:
                    has_any_shift = True
                    continue
                # 숫자가 발견되면 통계 행으로 판단
                try:
                    fv = float(dv)
                    if fv >= 0:  # 0 이상의 숫자 → 통계 행
                        is_stats_row = True
                        break
                except (ValueError, TypeError):
                    pass

            if is_stats_row:
                break  # 이 행부터는 통계 영역 → 이름 탐색 종료

            if has_any_shift:
                found_name_rows.append((r, c, sv))
                break  # 한 행당 하나의 이름만

    if not found_name_rows:
        raise ValueError("근무자 이름을 찾을 수 없습니다. 서식 파일을 확인해 주세요.")

    info['worker_row_start'] = found_name_rows[0][0]
    info['name_col'] = found_name_rows[0][1]

    # 1명당 행 수 계산 (2개 이상 이름이 있을 때)
    if len(found_name_rows) >= 2:
        step = found_name_rows[1][0] - found_name_rows[0][0]
        info['worker_row_step'] = max(1, step)
    else:
        info['worker_row_step'] = 1

    # ── 4. 실제 근무 기호가 있는 행 오프셋 탐지 ────────
    # 이름 행 자체 or 이름 행+1에서 shift 키워드 탐색
    step = info['worker_row_step']
    name_r = info['worker_row_start']
    data_offset = 0
    for offset in range(step):
        found_shift = False
        for c in range(info['date_col_start'], info['date_col_start'] + 10):
            v = cell_val(name_r + offset, c)
            if v in SHIFT_KEYWORDS:
                found_shift = True
                break
        if found_shift:
            data_offset = offset
            break
    info['worker_data_row_offset'] = data_offset

    # ── 5. 하단 통계(footer) 시작 행 탐지 ────────────────
    # found_name_rows의 마지막 이름 다음부터 footer 키워드 탐색
    last_worker_row = found_name_rows[-1][0] + step - 1
    for r in range(last_worker_row + 1, last_worker_row + 20):
        for c in range(1, SEARCH_MAX_COL + 1):
            v = cell_val(r, c)
            if v and str(v).strip() in FOOTER_KEYWORDS:
                info['footer_row'] = r
                break
        if info['footer_row']:
            break

    # ── 6. "n월" 셀 위치 탐지 ───────────────────────────
    MONTH_RE = _re.compile(r'\d{1,2}월')
    for r in range(1, search_start_row):
        for c in range(1, max_col + 1):
            v = cell_val(r, c)
            if v and isinstance(v, str) and MONTH_RE.search(v):
                info['month_cells'].append((r, c))

    # ── 7. 우측 통계 시작 열 탐지 ───────────────────────
    # 날짜 마지막 열 다음에 한글 헤더가 있으면 우측 통계로 간주
    if info['date_col_start']:
        last_date_col = info['date_col_start'] + info['template_days'] - 1
        for c in range(last_date_col + 1, last_date_col + 6):
            v = cell_val(info['date_row'] - 1, c) or cell_val(info['date_row'], c)
            if v and isinstance(v, str) and len(str(v).strip()) >= 1:
                info['right_stat_col_start'] = c
                break

    return info


def analyze_template_xw(ws_xw) -> dict:
    """
    xlwings sheet 객체를 받아 서식 구조를 분석하고 정보를 반환한다.
    (openpyxl의 IndexError를 피하기 위한 xlwings 전용 버전)
    """
    ws = ws_xw
    # used_range 기반으로 최대 범위 설정
    max_row = ws.used_range.last_cell.row
    max_col = ws.used_range.last_cell.column
    SEARCH_MAX_COL = min(max_col, 16)

    info = {
        'date_row': None, 'day_row': None,
        'date_col_start': None, 'name_col': None,
        'worker_row_start': None, 'worker_row_step': 1,
        'worker_data_row_offset': 0, 'footer_row': None,
        'month_cells': [], 'right_stat_col_start': None,
        'template_days': 31,
    }

    def cell_val(r, c):
        v = ws.range((r, c)).value
        if v is None: return None
        if isinstance(v, (int, float)): return v
        return str(v).strip()

    # 1. 날짜 행 탐색
    for r in range(1, 15):
        consecutive = 0
        prev = 0
        for c in range(1, SEARCH_MAX_COL + 1):
            v = cell_val(r, c)
            if isinstance(v, (int, float)) and int(v) == prev + 1:
                consecutive += 1
                if consecutive == 6: break
                prev = int(v)
            else:
                consecutive = 0
                prev = 0
        if consecutive >= 6:
            info['date_row'] = r
            for c in range(1, SEARCH_MAX_COL + 1):
                v = cell_val(r, c)
                if isinstance(v, (int, float)) and int(v) == 1:
                    info['date_col_start'] = c
                    break
            last_day = 0
            for c in range(info['date_col_start'], max_col + 1):
                v = cell_val(r, c)
                if isinstance(v, (int, float)) and int(v) > last_day:
                    last_day = int(v)
                elif last_day >= 28: break
            info['template_days'] = max(last_day, 28)
            break

    if not info['date_row']:
        raise ValueError("날짜 행을 찾을 수 없습니다.")

    # 2. 요일 행 탐색
    for r in range(max(1, info['date_row'] - 2), info['date_row'] + 3):
        if r == info['date_row']: continue
        day_cnt = 0
        for c in range(info['date_col_start'], info['date_col_start'] + 10):
            if cell_val(r, c) in KOREAN_DAYS: day_cnt += 1
        if day_cnt >= 5:
            info['day_row'] = r
            break

    # 3. 이름 및 근무자 시작 행 탐색
    import re as _re
    HANGUL_RE = _re.compile(r'^[\uAC00-\uD7A3]{2,}$')
    search_start = max(info['date_row'], (info['day_row'] or 0)) + 1
    found_names = []
    for r in range(search_start, min(search_start + 60, max_row + 1)):
        for c in range(1, 5):
            v = cell_val(r, c)
            if v and isinstance(v, str) and HANGUL_RE.match(v) and v not in FOOTER_KEYWORDS:
                # 숫자 통계 행인지 체크
                is_stats = False
                for dc in range(info['date_col_start'], info['date_col_start'] + 10):
                    dv = cell_val(r, dc)
                    if isinstance(dv, (int, float)):
                        is_stats = True; break
                if not is_stats:
                    found_names.append((r, c, v))
                    break
    
    if not found_names: raise ValueError("근무자 이름을 찾지 못했습니다.")
    info['worker_row_start'] = found_names[0][0]
    info['name_col'] = found_names[0][1]
    if len(found_names) >= 2:
        info['worker_row_step'] = max(1, found_names[1][0] - found_names[0][0])
    
    # 4. 데이터 오프셋 및 기타
    for offset in range(info['worker_row_step']):
        for c in range(info['date_col_start'], info['date_col_start'] + 5):
            if cell_val(info['worker_row_start'] + offset, c) in SHIFT_KEYWORDS:
                info['worker_data_row_offset'] = offset; break
    
    last_r = found_names[-1][0] + info['worker_row_step']
    for r in range(last_r, min(last_r + 20, max_row + 1)):
        for c in range(1, SEARCH_MAX_COL + 1):
            if cell_val(r, c) in FOOTER_KEYWORDS:
                info['footer_row'] = r; break
        if info['footer_row']: break

    MONTH_RE = _re.compile(r'\d{1,2}월')
    for r in range(1, search_start):
        for c in range(1, max_col + 1):
            v = cell_val(r, c)
            if v and isinstance(v, str) and MONTH_RE.search(v):
                info['month_cells'].append((r, c))

    if info['date_col_start']:
        last_c = info['date_col_start'] + info['template_days'] - 1
        for c in range(last_c + 1, last_c + 6):
            v = cell_val(info['date_row'], c) or cell_val(info['date_row'] - 1, c)
            if v and isinstance(v, str) and len(v) >= 1:
                info['right_stat_col_start'] = c; break

    return info


# ───────────────────────────────────────────────
# 근무자 명단 로드 (지능형 분석 기반)
# ───────────────────────────────────────────────
def create_dummy_workers(template_path: str = None) -> tuple[list[Worker], str | None]:
    """
    Returns (workers, error_message).
    error_message is None on success, or a string describing the issue.
    """
    workers = []
    error_msg = None

    if template_path and os.path.exists(template_path):
        # ── 1차 시도: 지능형 분석 엔진 ──────────────────
        try:
            import openpyxl
            wb = openpyxl.load_workbook(template_path, data_only=True)
            ws = wb.worksheets[0]
            info = analyze_template(ws)

            step = info['worker_row_step']
            r    = info['worker_row_start']
            nc   = info['name_col']
            while r <= ws.max_row:
                v = ws.cell(row=r, column=nc).value
                if v is None or str(v).strip() == "":
                    break
                sv = str(v).strip()
                if sv in FOOTER_KEYWORDS:
                    break
                workers.append(Worker(id=len(workers), name=sv))
                r += step
            wb.close()
        except Exception as e1:
            error_msg = f"[지능형 분석 실패, 기본 스캔으로 전환] {e1}"
            workers = []

        # ── 2차 폴백: 기존 B열 스캔 방식 ───────────────
        if not workers:
            try:
                import openpyxl
                wb2 = openpyxl.load_workbook(template_path, data_only=True)
                ws2 = wb2.worksheets[0]
                for row in range(3, 80):
                    name = ws2.cell(row=row, column=2).value  # B열
                    if name is None:
                        continue
                    sv = str(name).strip()
                    if not sv or sv in FOOTER_KEYWORDS:
                        if workers:  # 이름을 하나라도 찾은 뒤 빈칸이면 종료
                            break
                        continue
                    import re as _re2
                    if _re2.match(r'^[\uAC00-\uD7A3]{2,}$', sv):
                        workers.append(Worker(id=len(workers), name=sv))
                wb2.close()
                if workers:
                    error_msg = None  # 성공
            except Exception as e2:
                error_msg = f"기본 스캔도 실패: {e2}"
                workers = []

    if len(workers) < NUM_WORKERS:
        for i in range(len(workers), NUM_WORKERS):
            workers.append(Worker(id=i, name=""))

    return workers, error_msg


# ───────────────────────────────────────────────
# CSV 내보내기
# ───────────────────────────────────────────────
def export_to_csv(filepath: str, workers: list[Worker], schedule: list[list[int]], num_days: int):
    with open(filepath, mode='w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['이름'] + [f'{d+1}일' for d in range(num_days)])
        for w_idx, worker in enumerate(workers):
            writer.writerow([worker.name] + [SHIFT_TO_STR[schedule[w_idx][d]] for d in range(num_days)])


# ───────────────────────────────────────────────
# 수식 내 셀 참조 범위 보정 (예: C5:C12 → C5:C14)
# ───────────────────────────────────────────────
def _shift_formula_row(formula: str, old_last_row: int, new_last_row: int) -> str:
    """수식 안의 :C{old}  →  :C{new} 패턴을 모두 교체"""
    def replacer(m):
        col = m.group(1)
        row = int(m.group(2))
        if row == old_last_row:
            return f":{col}{new_last_row}"
        return m.group(0)
    return re.sub(r':([A-Z]+)(\d+)', replacer, formula)


def _apply_shift_colors_to_output_excel(
    output_path: str,
    workers: list[Worker],
    num_days: int,
    date_col_start: int,
    w_row_start: int,
    w_row_step: int,
    w_data_offset: int,
):
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(output_path)
    ws = wb.worksheets[0]
    for w_idx in range(len(workers)):
        data_r = w_row_start + w_idx * w_row_step + w_data_offset
        for d in range(num_days):
            cell = ws.cell(row=data_r, column=date_col_start + d)
            token = str(cell.value).strip() if cell.value is not None else "-"
            style = SHIFT_COLOR_RULES.get(token)
            if not style:
                continue

            new_font = copy(cell.font)
            new_font.color = style["font"]
            cell.font = new_font

            if style["fill"]:
                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color=style["fill"],
                    end_color=style["fill"],
                )
            else:
                cell.fill = PatternFill(fill_type=None)

    wb.save(output_path)
    wb.close()


def _with_replaced_border_side(border, side_name: str, side):
    from openpyxl.styles import Border

    values = {
        "left": copy(border.left),
        "right": copy(border.right),
        "top": copy(border.top),
        "bottom": copy(border.bottom),
        "diagonal": copy(border.diagonal),
        "diagonal_direction": border.diagonal_direction,
        "diagonalUp": border.diagonalUp,
        "diagonalDown": border.diagonalDown,
        "outline": border.outline,
        "vertical": copy(border.vertical),
        "horizontal": copy(border.horizontal),
    }
    values[side_name] = copy(side)
    return Border(**values)


def _normalize_worker_footer_separator(
    output_path: str,
    worker_row_start: int,
    worker_count: int,
    worker_row_step: int,
    footer_row: int | None,
) -> None:
    """근무자 수 증감 후 중간 굵은 구분선이 남지 않도록 내부선 기준으로 정리한다."""
    if not footer_row or worker_count <= 0:
        return

    import openpyxl

    wb = openpyxl.load_workbook(output_path)
    ws = wb.worksheets[0]
    last_worker_row = worker_row_start + worker_count * worker_row_step - 1
    if last_worker_row <= 1 or footer_row <= last_worker_row:
        wb.close()
        return

    source_row = max(worker_row_start, last_worker_row - worker_row_step)
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        source_bottom = ws.cell(source_row, col).border.bottom

        last_worker_cell = ws.cell(last_worker_row, col)
        last_worker_cell.border = _with_replaced_border_side(
            last_worker_cell.border,
            "bottom",
            source_bottom,
        )

        footer_cell = ws.cell(footer_row, col)
        footer_cell.border = _with_replaced_border_side(
            footer_cell.border,
            "top",
            source_bottom,
        )

    wb.save(output_path)
    wb.close()


def _is_basic_template_path(template_path: str) -> bool:
    basename = os.path.basename(template_path)
    return "기본 틀" in basename or basename == "default_template.xlsx"


def _set_cell_border_side(cell, side_name: str, side) -> None:
    cell.border = _with_replaced_border_side(cell.border, side_name, side)


def _apply_medium_border_line(ws, cells, side_name: str, side) -> None:
    for cell in cells:
        _set_cell_border_side(cell, side_name, side)


def _finalize_basic_template_layout(
    output_path: str,
    *,
    date_row: int,
    day_row: int | None,
    date_col_start: int,
    name_col: int,
    worker_row_start: int,
    worker_count: int,
    worker_row_step: int,
    worker_data_row_offset: int,
    num_days: int,
) -> None:
    """기본 틀 계열의 통계 수식과 테두리를 최종 구조 기준으로 재작성한다."""
    if worker_count <= 0 or num_days <= 0:
        return

    import openpyxl
    from openpyxl.styles import Side

    wb = openpyxl.load_workbook(output_path)
    ws = wb.worksheets[0]

    label_row = day_row or date_row
    last_date_col = date_col_start + num_days - 1
    right_stat_col = last_date_col + 1
    stat_cols = {
        "주": right_stat_col,
        "야": right_stat_col + 1,
        "비": right_stat_col + 2,
        "휴": right_stat_col + 3,
        "연": right_stat_col + 4,
        "근무시간": right_stat_col + 5,
    }
    time_col = stat_cols["근무시간"]
    footer_row = worker_row_start + worker_count * worker_row_step
    footer_labels = [("주간", "주"), ("야간", "야"), ("비번", "비"), ("휴무", "휴")]
    last_footer_row = footer_row + len(footer_labels) - 1
    first_worker_data_row = worker_row_start + worker_data_row_offset
    last_worker_data_row = (
        worker_row_start + (worker_count - 1) * worker_row_step + worker_data_row_offset
    )
    date_start_letter = get_column_letter(date_col_start)
    date_end_letter = get_column_letter(last_date_col)

    for label, col in stat_cols.items():
        ws.cell(label_row, col).value = label

    for w_idx in range(worker_count):
        data_row = worker_row_start + w_idx * worker_row_step + worker_data_row_offset
        row_range = f"{date_start_letter}{data_row}:{date_end_letter}{data_row}"
        ws.cell(data_row, stat_cols["주"]).value = f'=COUNTIF({row_range},"주")'
        ws.cell(data_row, stat_cols["야"]).value = f'=COUNTIF({row_range},"야")'
        ws.cell(data_row, stat_cols["비"]).value = f'=COUNTIF({row_range},"비")'
        ws.cell(data_row, stat_cols["휴"]).value = f'=COUNTIF({row_range},"휴")'
        ws.cell(data_row, stat_cols["연"]).value = f'=COUNTIF({row_range},"연")'
        row_refs = [
            f'{get_column_letter(stat_cols["주"])}{data_row}',
            f'{get_column_letter(stat_cols["연"])}{data_row}',
            f'{get_column_letter(stat_cols["야"])}{data_row}',
            f'{get_column_letter(stat_cols["비"])}{data_row}',
        ]
        ws.cell(data_row, stat_cols["근무시간"]).value = f'=({" + ".join(row_refs)})*8'

    for offset, (label, shift_token) in enumerate(footer_labels):
        row = footer_row + offset
        ws.cell(row, name_col).value = label
        for day_idx in range(num_days):
            col = date_col_start + day_idx
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'=COUNTIF({col_letter}{first_worker_data_row}:'
                f'{col_letter}{last_worker_data_row},"{shift_token}")'
            )

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    for row in range(date_row, last_footer_row + 1):
        for col in range(name_col, time_col + 1):
            cell = ws.cell(row, col)
            cell.border = _with_replaced_border_side(cell.border, "top", thin)
            cell.border = _with_replaced_border_side(cell.border, "bottom", thin)
            cell.border = _with_replaced_border_side(cell.border, "left", thin)
            cell.border = _with_replaced_border_side(cell.border, "right", thin)

    # 전체 표 외곽
    _apply_medium_border_line(
        ws,
        [ws.cell(date_row, col) for col in range(name_col, last_date_col + 1)]
        + [ws.cell(label_row, col) for col in range(right_stat_col, time_col + 1)],
        "top",
        medium,
    )
    _apply_medium_border_line(
        ws,
        [ws.cell(last_footer_row, col) for col in range(name_col, last_date_col + 1)],
        "bottom",
        medium,
    )
    _apply_medium_border_line(
        ws,
        [ws.cell(row, name_col) for row in range(date_row, last_footer_row + 1)],
        "left",
        medium,
    )
    _apply_medium_border_line(
        ws,
        [ws.cell(row, time_col) for row in range(label_row, last_worker_data_row + 1)],
        "right",
        medium,
    )

    # 날짜 영역과 우측 개인 통계 영역의 경계
    _apply_medium_border_line(
        ws,
        [ws.cell(row, last_date_col) for row in range(date_row, last_footer_row + 1)],
        "right",
        medium,
    )
    _apply_medium_border_line(
        ws,
        [ws.cell(row, right_stat_col) for row in range(label_row, last_worker_data_row + 1)],
        "left",
        medium,
    )

    # 근무자 영역과 하단 일자별 통계 영역의 경계
    last_worker_row = worker_row_start + (worker_count - 1) * worker_row_step
    _apply_medium_border_line(
        ws,
        [ws.cell(last_worker_row, col) for col in range(name_col, time_col + 1)],
        "bottom",
        medium,
    )
    _apply_medium_border_line(
        ws,
        [ws.cell(footer_row, col) for col in range(name_col, last_date_col + 1)],
        "top",
        medium,
    )

    # 하단 통계는 날짜 영역까지만 사용하므로 우측 개인 통계 영역과 분리한다.
    for row in range(footer_row, last_footer_row + 1):
        for col in range(right_stat_col, time_col + 1):
            ws.cell(row, col).value = None
            ws.cell(row, col).border = _with_replaced_border_side(ws.cell(row, col).border, "top", Side())
            ws.cell(row, col).border = _with_replaced_border_side(ws.cell(row, col).border, "bottom", Side())
            ws.cell(row, col).border = _with_replaced_border_side(ws.cell(row, col).border, "left", Side())
            ws.cell(row, col).border = _with_replaced_border_side(ws.cell(row, col).border, "right", Side())

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        pass

    wb.save(output_path)
    wb.close()


# ───────────────────────────────────────────────
# 엑셀 지능형 출력 (핵심 함수)
# ───────────────────────────────────────────────
def export_to_excel(template_path: str, output_path: str,
                    workers: list[Worker], schedule: list[list[int]],
                    num_days: int, year: int, month: int,
                    apply_shift_colors: bool = False):
    try:
        import xlwings as xw
        import openpyxl
    except ImportError as e:
        raise ImportError(f"필요한 라이브러리가 없습니다: {e}")

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"서식 파일({template_path})을 찾을 수 없습니다.")

    # ── 분석: openpyxl로 구조를 먼저 파악하고, 실제 편집은 Excel 엔진에 맡긴다. ──
    try:
        wb_tmp = openpyxl.load_workbook(template_path, data_only=False)
        try:
            ws_tmp = wb_tmp.worksheets[0]
            info = analyze_template_v3(ws_tmp)
            merged_ranges = _read_merged_ranges(ws_tmp)
        finally:
            wb_tmp.close()
    except Exception as e:
        raise RuntimeError(f"서식 분석 중 오류가 발생했습니다: {e}")

    date_row       = info['date_row']
    day_row        = info['day_row']
    date_col_start = info['date_col_start']
    name_col       = info['name_col']
    w_row_start    = info['worker_row_start']
    w_row_step     = info['worker_row_step']
    w_data_offset  = info['worker_data_row_offset']
    footer_row_org = info['footer_row']
    tmpl_days      = info['template_days']
    right_stat_col = info['right_stat_col_start']

    # ── xlwings 로 실제 편집 ────────────────────
    xw_app = xw.App(visible=False)
    try:
        wb = xw_app.books.open(os.path.abspath(template_path))
        ws = wb.sheets[0]
        _unmerge_ranges_xw(ws, merged_ranges)

        korean_days = ["월", "화", "수", "목", "금", "토", "일"]

        # ── 0. 날짜 열 구조 조정 ─────────────────────────
        if num_days > tmpl_days:
            extra_days = num_days - tmpl_days
            current_days = tmpl_days
            for _ in range(extra_days):
                insert_col = _middle_date_col(date_col_start, current_days)
                source_col = insert_col - 1
                ws.api.Columns(insert_col).Insert()
                merged_ranges = _adjust_merged_ranges_for_col_insert(merged_ranges, insert_col)
                ws.range((1, source_col), (ws.used_range.last_cell.row, source_col)).copy(
                    ws.range((1, insert_col))
                )
                current_days += 1
            tmpl_days = num_days
            right_stat_col = right_stat_col + extra_days if right_stat_col else None
        elif num_days < tmpl_days:
            remove_days = tmpl_days - num_days
            current_days = tmpl_days
            for _ in range(remove_days):
                col_to_del = _middle_date_col(date_col_start, current_days)
                ws.api.Columns(col_to_del).Delete()
                merged_ranges = _adjust_merged_ranges_for_col_delete(merged_ranges, col_to_del)
                current_days -= 1
            tmpl_days = num_days
            right_stat_col = right_stat_col - remove_days if right_stat_col else None

        # ── 1. n년 n월 / n월 텍스트 교체 ──────────────────
        for (r, c) in info['month_cells']:
            old_val = str(ws.range((r, c)).value or "")
            new_val = re.sub(r'\d{4}\s*년\s*\d{1,2}\s*월', f"{year}년 {month}월", old_val)
            new_val = re.sub(r'\d{1,2}\s*월', f"{month}월", new_val)
            ws.range((r, c)).value = new_val

        # ── 2. 날짜 / 요일 헤더 갱신 (31칸 기준 먼저 쓰고 나머지는 나중에 삭제) ──
        for d in range(tmpl_days):
            col = date_col_start + d
            if d < num_days:
                ws.range((date_row, col)).value = d + 1
                if day_row:
                    ws.range((day_row, col)).value = korean_days[
                        datetime.date(year, month, d + 1).weekday()
                    ]
            else:
                ws.range((date_row, col)).value = ""
                if day_row:
                    ws.range((day_row, col)).value = ""

        # ── 3. 기존 근무자 수 파악 ──────────────
        base_workers = 0
        r = w_row_start
        while True:
            v = ws.range((r, name_col)).value
            sv = str(v).strip() if v else ""
            if not sv or sv in FOOTER_KEYWORDS:
                break
            base_workers += 1
            r += w_row_step

        # ── 4. 인원 행 삽입 / 삭제 ──────────────
        diff = len(workers) - base_workers
        footer_row_new = footer_row_org + diff * w_row_step if footer_row_org else None
        if diff > 0:
            # 인원 추가: 마지막 행의 하단 테두리를 보존하기 위해 근무자 영역 중간에 삽입
            insert_at = _middle_worker_insert_row(w_row_start, base_workers, w_row_step)
            src_start = max(w_row_start, insert_at - w_row_step)
            for _ in range(diff):
                ws.api.Rows(f"{insert_at}:{insert_at + w_row_step - 1}").Insert()
                merged_ranges = _adjust_merged_ranges_for_row_insert(
                    merged_ranges,
                    insert_at,
                    w_row_step,
                )
                # 중간 근무자 블록의 서식/수식을 복사해 마지막 행 전용 테두리 복제를 피한다.
                dst_start = insert_at
                for step_r in range(w_row_step):
                    ws.range((src_start + step_r, 1),
                              (src_start + step_r, ws.used_range.last_cell.column)).copy(
                        ws.range((dst_start + step_r, 1))
                    )
                insert_at += w_row_step

        elif diff < 0:
            # 인원 감소: 마지막 행 전용 테두리를 보존하기 위해 근무자 영역 중간에서 삭제
            current_workers = base_workers
            for _ in range(-diff):
                del_start = _middle_worker_delete_row(w_row_start, current_workers, w_row_step)
                ws.api.Rows(f"{del_start}:{del_start + w_row_step - 1}").Delete()
                merged_ranges = _adjust_merged_ranges_for_row_delete(
                    merged_ranges,
                    del_start,
                    w_row_step,
                )
                current_workers -= 1

        # ── 5. 근무자 이름 / 데이터 쓰기 ─────────
        for w_idx, worker in enumerate(workers):
            name_r = w_row_start + w_idx * w_row_step
            data_r = name_r + w_data_offset
            ws.range((name_r, name_col)).value = worker.name
            for d in range(num_days):
                    s_val = schedule[w_idx][d]
                    if isinstance(s_val, int):
                        ws.range((data_r, date_col_start + d)).value = SHIFT_TO_STR.get(s_val, str(s_val))
                    else:
                        # '교', '출' 등 커스텀 약어는 그대로 입력
                        ws.range((data_r, date_col_start + d)).value = s_val
            # 나머지 날짜 칸 비우기
            for d in range(num_days, tmpl_days):
                ws.range((data_r, date_col_start + d)).value = ""

        # ── 6. 하단 통계 수식 범위 보정 ──────────
        last_worker_row_new = w_row_start + len(workers) * w_row_step - 1
        last_worker_row_old = w_row_start + base_workers * w_row_step - 1

        if footer_row_new and diff != 0:
            # 통계 수식에서 참조범위의 마지막 행 번호를 교체
            stat_row = footer_row_new
            for r in range(stat_row, stat_row + 15):
                for c in range(date_col_start, date_col_start + tmpl_days):
                    cell = ws.range((r, c))
                    f = cell.formula
                    if f and f.startswith("=") and str(last_worker_row_old) in f:
                        cell.formula = _shift_formula_row(f, last_worker_row_old, last_worker_row_new)

        # ── 7. 우측 통계 수식 범위 보정 ──────────
        if right_stat_col:
            for w_idx in range(len(workers)):
                data_r = w_row_start + w_idx * w_row_step + w_data_offset
                if data_r == w_row_start + w_data_offset:
                    continue  # 첫 행은 원본 수식 그대로 사용
                # 첫 번째 근무자 우측 통계 수식을 복사
                src_r = w_row_start + w_data_offset
                for c_offset in range(8):  # 최대 8개 열 확인
                    c = right_stat_col + c_offset
                    src_formula = ws.range((src_r, c)).formula
                    if not src_formula or not src_formula.startswith("="):
                        break
                    ws.range((src_r, c)).copy(ws.range((data_r, c)))

        # ── 9. 마무리 저장 ────────────────────────
        _remerge_ranges_xw(ws, merged_ranges)
        ws.range((date_row, name_col)).select()
        wb.save(os.path.abspath(output_path))
        wb.close()
    finally:
        xw_app.quit()

    if _is_basic_template_path(template_path):
        _finalize_basic_template_layout(
            output_path=output_path,
            date_row=date_row,
            day_row=day_row,
            date_col_start=date_col_start,
            name_col=name_col,
            worker_row_start=w_row_start,
            worker_count=len(workers),
            worker_row_step=w_row_step,
            worker_data_row_offset=w_data_offset,
            num_days=num_days,
        )
    elif diff != 0:
        _normalize_worker_footer_separator(
            output_path=output_path,
            worker_row_start=w_row_start,
            worker_count=len(workers),
            worker_row_step=w_row_step,
            footer_row=footer_row_new,
        )

    # 템플릿의 색상과 서식은 항상 보존한다. apply_shift_colors는 V3에서 무시한다.
