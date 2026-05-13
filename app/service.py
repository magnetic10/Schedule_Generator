from __future__ import annotations

import calendar
import datetime
import math
import re
from dataclasses import replace
from pathlib import Path
from typing import Iterable

import holidays

from core.config import (
    NUM_WORKERS,
    SHIFT_DAY,
    SHIFT_LEAVE,
    SHIFT_NIGHT,
    SHIFT_OFF,
    SHIFT_OFF_NIGHT,
    SHIFT_TO_STR,
)
from core.data_io import (
    FOOTER_KEYWORDS,
    HEADER_KEYWORDS,
    analyze_template_v3,
    export_to_excel,
    is_actual_worker_name,
)
from core.models import Worker
from core.scheduler import SchedulerError, repair_schedule, solve_schedule

from .schemas import (
    ExcelExportResult,
    LeaveNeedGuide,
    MonthInfo,
    RepairEdit,
    ScheduleRepairRequest,
    ScheduleRequest,
    ScheduleResult,
    ScheduleSettings,
    TemplateWorkerLoadResult,
    SidebarDefaults,
    WorkerInput,
    WorkerScheduleRow,
)


SHIFT_CODE_TO_VALUE = {
    "day": SHIFT_DAY,
    "night": SHIFT_NIGHT,
    "off_night": SHIFT_OFF_NIGHT,
    "off": SHIFT_OFF,
    "leave": SHIFT_LEAVE,
}

FIXED_SHIFT_CODES = {"day", "night", "off_night", "off", "leave"}
SHIFT_VALUE_TO_CODE = {value: code for code, value in SHIFT_CODE_TO_VALUE.items()}
SHIFT_LABEL_TO_VALUE = {label: value for value, label in SHIFT_TO_STR.items()}

TEMPLATE_INFO_KEYS = (
    "date_row",
    "day_row",
    "date_col_start",
    "name_col",
    "worker_row_start",
    "worker_row_step",
    "worker_data_row_offset",
    "footer_row",
    "right_stat_col_start",
    "template_days",
    "sheet_name",
    "worker_row_count",
    "has_real_worker_names",
)


def next_month_from(today: datetime.date | None = None) -> tuple[int, int]:
    base = today or datetime.date.today()
    if base.month == 12:
        return base.year + 1, 1
    return base.year, base.month + 1


def build_month_info(year: int, month: int) -> MonthInfo:
    if not 2000 <= int(year) <= 2100:
        raise ValueError("연도는 2000부터 2100 사이여야 합니다.")
    if not 1 <= int(month) <= 12:
        raise ValueError("월은 1부터 12 사이여야 합니다.")

    days_in_month = calendar.monthrange(year, month)[1]
    kr_holidays = holidays.KR(years=year)
    workday_count = 0

    for day in range(1, days_in_month + 1):
        current = datetime.date(year, month, day)
        if current.weekday() < 5 and current not in kr_holidays:
            workday_count += 1

    return MonthInfo(
        year=year,
        month=month,
        days_in_month=days_in_month,
        workday_count=workday_count,
        default_target_hours=workday_count * 8,
        label=f"{year}년 {month}월 | 총 {days_in_month}일 | 평일 {workday_count}일",
    )


def build_sidebar_defaults(year: int | None = None, month: int | None = None) -> SidebarDefaults:
    if year is None or month is None:
        year, month = next_month_from()
    month_info = build_month_info(year, month)

    return SidebarDefaults(
        month_info=month_info,
        common_target_hours=month_info.default_target_hours,
        settings=ScheduleSettings(),
    )


def reset_worker_schedules(worker_inputs: Iterable[WorkerInput]) -> list[WorkerInput]:
    return [replace(worker, fixed_shifts={}) for worker in worker_inputs]


def build_solver_config(req: ScheduleRequest) -> dict:
    settings = req.settings
    cfg = {
        "tgt_d": settings.target_day,
        "tgt_n": settings.target_night,
        "allow_min_d": settings.min_day,
        "allow_max_d": settings.max_day,
        "allow_min_n": settings.min_night,
        "allow_max_n": settings.max_night,
        "use_emergency_range": settings.use_emergency_range,
        "special_shifts": settings.special_shifts,
        "use_preference": settings.use_preference,
        "allow_leave_after_off_night": settings.allow_leave_after_off_night,
        "allow_double_night_cycle": settings.allow_double_night_cycle,
    }

    if settings.use_emergency_range:
        cfg.update(
            {
                "em_min_d": settings.emergency_min_day
                if settings.emergency_min_day is not None
                else settings.min_day,
                "em_max_d": settings.emergency_max_day
                if settings.emergency_max_day is not None
                else settings.max_day,
                "em_min_n": settings.emergency_min_night
                if settings.emergency_min_night is not None
                else settings.min_night,
                "em_max_n": settings.emergency_max_night
                if settings.emergency_max_night is not None
                else settings.max_night,
            }
        )

    return cfg


def build_workers(worker_inputs: Iterable[WorkerInput], days_in_month: int) -> list[Worker]:
    workers: list[Worker] = []
    for idx, item in enumerate(worker_inputs):
        name = item.name.strip()
        worker = Worker(id=idx, name=name)
        dedicated_shift = _dedicated_shift_from_input(item)
        worker.dedicated_shift = dedicated_shift or ""
        worker.is_only_day = dedicated_shift == "day"
        worker.prev_month_last_day_night = bool(item.prev_month_last_day_night)
        worker.target_hours = 160 if item.target_hours is None else int(item.target_hours)
        worker.start_day = max(0, int(item.start_day) - 1)
        end_day = item.end_day if item.end_day is not None else days_in_month
        worker.end_day = min(days_in_month - 1, int(end_day) - 1)

        if item.preference in ("day", "night"):
            worker.shift_preference = item.preference
        else:
            worker.shift_preference = None

        for day_number, shift in item.fixed_shifts.items():
            day_index = int(day_number) - 1
            if day_index < 0 or day_index >= days_in_month:
                raise SchedulerError(
                    f"{worker.name}: fixed_shifts day {day_number} is out of range."
                )

            if shift == "day":
                worker.fixed_day_days.append(day_index)
            elif shift == "night":
                worker.fixed_night_days.append(day_index)
            elif shift == "off_night":
                worker.fixed_off_night_days.append(day_index)
            elif shift == "off":
                worker.fixed_off_days.append(day_index)
            elif shift == "leave":
                worker.fixed_leave_days.append(day_index)
            else:
                token = str(shift).strip()
                if not token:
                    continue
                worker.custom_shifts.setdefault(token, []).append(day_index)

        workers.append(worker)

    return workers


def _dedicated_shift_from_input(item: WorkerInput) -> str | None:
    value = (item.dedicated_shift or "").strip()
    if value not in ("", "day", "night"):
        raise SchedulerError(f"{item.name}: 전담 근무 값이 올바르지 않습니다.")
    if not value and item.is_day_only:
        value = "day"
    return value or None


def build_default_worker_inputs(count: int = NUM_WORKERS) -> list[WorkerInput]:
    return [WorkerInput(name="") for _ in range(count)]


def load_template_workers(
    template_path: str | None,
    default_count: int = NUM_WORKERS,
) -> TemplateWorkerLoadResult:
    if not template_path:
        return TemplateWorkerLoadResult(
            workers=build_default_worker_inputs(default_count),
            recognized_count=0,
        )

    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(f"엑셀 서식 파일을 찾을 수 없습니다: {template_path}")

    names, warning, template_info = _read_worker_names_from_template(path)
    if names:
        return TemplateWorkerLoadResult(
            workers=[WorkerInput(name=name) for name in names],
            recognized_count=len(names),
            warning=warning,
            template_info=template_info,
        )

    if template_info.get("date_row") and template_info.get("worker_row_start"):
        return TemplateWorkerLoadResult(
            workers=[],
            recognized_count=0,
            warning=warning or "서식은 불러왔지만 실제 근무자 이름을 찾지 못해 현재 입력표를 유지합니다.",
            template_info=template_info,
            template_status="valid_without_names",
            preserve_existing_workers=True,
        )

    return TemplateWorkerLoadResult(
        workers=build_default_worker_inputs(default_count),
        recognized_count=0,
        warning=warning or "업로드 파일에서 근무자 이름을 찾지 못해 기본 빈 행을 사용합니다.",
        template_info=template_info,
    )


def _read_worker_names_from_template(path: Path) -> tuple[list[str], str | None, dict[str, object]]:
    warning: str | None = None
    template_info: dict[str, object] = {}

    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            ws = wb.worksheets[0]
            info = analyze_template_v3(ws)
            template_info = _public_template_info(info)
            names = _extract_worker_names_from_template_info(ws, info)
            if names:
                return names, None, template_info
            warning = "서식은 불러왔지만 실제 근무자 이름을 찾지 못했습니다."
            return [], warning, template_info
        finally:
            wb.close()
    except Exception as exc:
        warning = f"서식 분석 실패, 기본 스캔으로 전환했습니다: {exc}"

    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            ws = wb.worksheets[0]
            names = _scan_worker_names_from_b_column(ws)
            if names:
                return names, warning, template_info
        finally:
            wb.close()
    except Exception as exc:
        if warning:
            warning = f"{warning} 기본 스캔도 실패했습니다: {exc}"
        else:
            warning = f"기본 스캔 실패: {exc}"

    return [], warning or "업로드 파일에서 근무자 이름을 찾지 못했습니다.", template_info


def _extract_worker_names_from_template_info(ws, info: dict) -> list[str]:
    names: list[str] = []
    row = int(info["worker_row_start"])
    step = max(1, int(info.get("worker_row_step") or 1))
    name_col = int(info["name_col"])

    while row <= ws.max_row:
        raw_name = ws.cell(row=row, column=name_col).value
        if raw_name is None:
            break
        name = str(raw_name).strip()
        if not name or name in FOOTER_KEYWORDS or name in HEADER_KEYWORDS:
            break
        if is_actual_worker_name(name):
            names.append(name)
        row += step

    return names


def _scan_worker_names_from_b_column(ws) -> list[str]:
    names: list[str] = []
    hangul_name_re = re.compile(r"^[\uAC00-\uD7A3]{2,}$")

    for row in range(3, 80):
        raw_name = ws.cell(row=row, column=2).value
        if raw_name is None:
            continue
        name = str(raw_name).strip()
        if not name or name in FOOTER_KEYWORDS or name in HEADER_KEYWORDS:
            if names:
                break
            continue
        if hangul_name_re.match(name) and is_actual_worker_name(name):
            names.append(name)

    return names


def _public_template_info(info: dict) -> dict[str, object]:
    return {key: info.get(key) for key in TEMPLATE_INFO_KEYS}


def solve_request(req: ScheduleRequest) -> ScheduleResult:
    days_in_month = calendar.monthrange(req.year, req.month)[1]
    _validate_pre_solver_conditions(req, days_in_month)
    workers = build_workers(req.workers, days_in_month)
    cfg = build_solver_config(req)
    schedule, status, solver_info = solve_schedule(
        workers,
        days_in_month,
        cfg,
        random_seed=req.random_seed,
    )

    return _schedule_result_from_solver(
        req=req,
        workers=workers,
        schedule=schedule,
        status=status,
        solver_info=solver_info,
    )


def repair_request(req: ScheduleRepairRequest) -> ScheduleResult:
    days_in_month = calendar.monthrange(req.request.year, req.request.month)[1]
    _validate_pre_solver_conditions(req.request, days_in_month)
    _validate_repair_shape(req, days_in_month)

    workers = build_workers(req.request.workers, days_in_month)
    cfg = build_solver_config(req.request)
    baseline_schedule = [
        [_normalize_result_shift(value) for value in row.raw_days]
        for row in req.result.rows
    ]
    locked_assignments = _build_repair_locks(req.edits, req.request, baseline_schedule, days_in_month)

    schedule, status, solver_info = repair_schedule(
        workers,
        days_in_month,
        cfg,
        baseline_schedule=baseline_schedule,
        locked_assignments=locked_assignments,
        random_seed=req.request.random_seed,
    )

    return _schedule_result_from_solver(
        req=req.request,
        workers=workers,
        schedule=schedule,
        status=status,
        solver_info=solver_info,
    )


def _schedule_result_from_solver(
    req: ScheduleRequest,
    workers: list[Worker],
    schedule: list[list[int | str]],
    status: str,
    solver_info: dict,
) -> ScheduleResult:
    days_in_month = calendar.monthrange(req.year, req.month)[1]
    rows = [
        WorkerScheduleRow(
            name=display_worker_name(worker.name, idx),
            raw_days=list(schedule[idx]),
            days=[
                SHIFT_TO_STR.get(value, str(value)) if isinstance(value, int) else value
                for value in schedule[idx]
            ],
        )
        for idx, worker in enumerate(workers)
    ]

    day_counts = [
        sum(1 for row in schedule if row[day] == SHIFT_DAY)
        for day in range(days_in_month)
    ]
    night_counts = [
        sum(1 for row in schedule if row[day] == SHIFT_NIGHT)
        for day in range(days_in_month)
    ]

    return ScheduleResult(
        year=req.year,
        month=req.month,
        days_in_month=days_in_month,
        status=status,
        rows=rows,
        day_counts=day_counts,
        night_counts=night_counts,
        emergency_days=list(solver_info.get("emergency_days", [])),
        double_night_cycle_days=list(solver_info.get("double_night_cycle_days", [])),
        double_night_cycle_used=bool(solver_info.get("double_night_cycle_used", False)),
        long_off_streak_fallback_used=bool(solver_info.get("long_off_streak_fallback_used", False)),
        long_off_streaks=_find_long_off_streaks(schedule, workers),
        repair_changed_count=int(solver_info.get("repair_changed_count", 0) or 0),
        repair_changed_cells=list(solver_info.get("repair_changed_cells", [])),
    )


def _validate_repair_shape(req: ScheduleRepairRequest, days_in_month: int) -> None:
    if req.result.year != req.request.year or req.result.month != req.request.month:
        raise SchedulerError("부분 재생성 기준 결과의 연도/월이 현재 입력 조건과 다릅니다.")
    if req.result.days_in_month != days_in_month:
        raise SchedulerError("부분 재생성 기준 결과의 날짜 수가 현재 월과 다릅니다.")
    if len(req.result.rows) != len(req.request.workers):
        raise SchedulerError("부분 재생성 기준 결과의 근무자 수가 현재 입력표와 다릅니다.")
    for row in req.result.rows:
        if len(row.raw_days) != days_in_month:
            raise SchedulerError("부분 재생성 기준 결과의 날짜별 값이 현재 월과 맞지 않습니다.")


def _build_repair_locks(
    edits: list[RepairEdit],
    req: ScheduleRequest,
    baseline_schedule: list[list[int | str]],
    days_in_month: int,
) -> dict[tuple[int, int], int | str]:
    if not edits:
        raise SchedulerError("부분 편집된 칸이 없습니다.")

    locks: dict[tuple[int, int], int | str] = {}
    for edit in edits:
        worker_index = int(edit.worker_index)
        day_number = int(edit.day)
        if worker_index < 0 or worker_index >= len(req.workers):
            raise SchedulerError("부분 편집한 근무자 행이 현재 입력표 범위를 벗어났습니다.")
        if day_number < 1 or day_number > days_in_month:
            raise SchedulerError("부분 편집한 날짜가 현재 월 범위를 벗어났습니다.")

        worker = req.workers[worker_index]
        if day_number < int(worker.start_day or 1):
            raise SchedulerError(f"{display_worker_name(worker.name, worker_index)}: {day_number}일은 근무 시작일 이전이라 편집할 수 없습니다.")
        if worker.end_day is not None and day_number > int(worker.end_day):
            raise SchedulerError(f"{display_worker_name(worker.name, worker_index)}: {day_number}일은 근무 종료일 이후라 편집할 수 없습니다.")

        shift = _shift_input_to_solver_value(edit.shift, req.settings.special_shifts)
        dedicated_shift = _dedicated_shift_from_input(worker)
        if dedicated_shift == "day" and shift in (SHIFT_NIGHT, SHIFT_OFF_NIGHT):
            raise SchedulerError(
                f"{display_worker_name(worker.name, worker_index)}: 주간 전담 근무자는 {day_number}일을 야간/비번으로 부분 편집할 수 없습니다."
            )
        if dedicated_shift == "night" and shift == SHIFT_DAY:
            raise SchedulerError(
                f"{display_worker_name(worker.name, worker_index)}: 야간 전담 근무자는 {day_number}일을 주간으로 부분 편집할 수 없습니다."
            )
        fixed_shift = _fixed_shift_for_day(worker, day_number)
        if fixed_shift:
            fixed_value = _shift_input_to_solver_value(fixed_shift, req.settings.special_shifts)
            if fixed_value != shift:
                raise SchedulerError(
                    f"{display_worker_name(worker.name, worker_index)}: {day_number}일은 입력표에서 고정한 근무라 결과에서 변경할 수 없습니다."
                )

        key = (worker_index, day_number - 1)
        if key in locks and locks[key] != shift:
            raise SchedulerError("같은 결과 칸에 서로 다른 부분 편집 값이 중복되었습니다.")
        locks[key] = shift

    return locks


def _shift_input_to_solver_value(shift: str | int, special_shifts: dict[str, int]) -> int | str:
    if isinstance(shift, int):
        if shift in SHIFT_VALUE_TO_CODE:
            return shift
        raise SchedulerError(f"알 수 없는 근무 코드입니다: {shift}")
    token = str(shift or "").strip()
    if not token:
        raise SchedulerError("부분 편집 값이 비어 있습니다.")
    if token in SHIFT_CODE_TO_VALUE:
        return SHIFT_CODE_TO_VALUE[token]
    if token in SHIFT_TO_STR.values():
        value = SHIFT_LABEL_TO_VALUE.get(token)
        if value is not None:
            return value
    if token in special_shifts:
        return token
    raise SchedulerError(f"알 수 없는 근무 값입니다: {token}")


def _normalize_result_shift(value: int | str) -> int | str:
    if isinstance(value, int):
        return value
    token = str(value or "").strip()
    if not token:
        return ""
    if token in SHIFT_CODE_TO_VALUE:
        return SHIFT_CODE_TO_VALUE[token]
    if token in SHIFT_TO_STR.values():
        value = SHIFT_LABEL_TO_VALUE.get(token)
        return value if value is not None else token
    return token


def _find_long_off_streaks(schedule: list[list[int | str]], workers: list[Worker]) -> list[dict[str, object]]:
    streaks: list[dict[str, object]] = []
    rest_values = {SHIFT_OFF, SHIFT_LEAVE}
    for worker_index, row in enumerate(schedule):
        start: int | None = None
        fixed_off_days = set(workers[worker_index].fixed_off_days)
        fixed_leave_days = set(workers[worker_index].fixed_leave_days)
        fixed_rest_days = fixed_off_days | fixed_leave_days

        for day_index in range(len(row) + 1):
            is_rest = day_index < len(row) and row[day_index] in rest_values
            if is_rest:
                if start is None:
                    start = day_index
                continue

            if start is not None:
                end = day_index - 1
                length = end - start + 1
                if length >= 5:
                    fixed_days = [day + 1 for day in range(start, end + 1) if day in fixed_rest_days]
                    fixed_leave_day_numbers = [
                        day + 1 for day in range(start, end + 1) if day in fixed_leave_days
                    ]
                    auto_days = [day + 1 for day in range(start, end + 1) if day not in fixed_rest_days]
                    streaks.append(
                        {
                            "worker_index": worker_index,
                            "worker_name": display_worker_name(workers[worker_index].name, worker_index),
                            "start_day": start + 1,
                            "end_day": end + 1,
                            "length": length,
                            "fixed_days": fixed_days,
                            "fixed_leave_days": fixed_leave_day_numbers,
                            "auto_days": auto_days,
                        }
                    )
                start = None

    return streaks


def build_leave_need_guide(req: ScheduleRequest) -> LeaveNeedGuide:
    days_in_month = calendar.monthrange(req.year, req.month)[1]
    _validate_pre_solver_conditions(req, days_in_month)
    total_target_hours = _total_target_hours(req.workers)
    leave_credit_hours, special_credit_hours = _non_to_credit_hours(req)
    non_to_credit_hours = leave_credit_hours + special_credit_hours
    remaining_regular_target_hours = max(0, total_target_hours - non_to_credit_hours)
    max_regular_capacity_hours = _max_regular_capacity_hours(req.settings, days_in_month)
    shortage_hours = max(0, remaining_regular_target_hours - max_regular_capacity_hours)
    suggested_leave_days = math.ceil(shortage_hours / 8) if shortage_hours else 0
    uses_individual_targets = any(worker.target_hours is not None for worker in req.workers)
    uses_emergency_range = bool(req.settings.use_emergency_range)
    capacity_basis_label = "예외 범위 포함" if uses_emergency_range else "기본 범위"

    if shortage_hours:
        message = (
            f"현재 목표시간 기준으로 {capacity_basis_label} 최대 배치 가능 시간보다 {shortage_hours}시간 많습니다. "
            f"연가 {suggested_leave_days}일이 더 필요합니다."
        )
    else:
        message = f"현재 목표시간과 입력된 연가 기준으로 {capacity_basis_label} 최대 배치 가능 시간 안에 들어옵니다."

    return LeaveNeedGuide(
        year=req.year,
        month=req.month,
        days_in_month=days_in_month,
        total_target_hours=total_target_hours,
        leave_credit_hours=leave_credit_hours,
        special_credit_hours=special_credit_hours,
        non_to_credit_hours=non_to_credit_hours,
        remaining_regular_target_hours=remaining_regular_target_hours,
        max_regular_capacity_hours=max_regular_capacity_hours,
        shortage_hours=shortage_hours,
        suggested_leave_days=suggested_leave_days,
        uses_individual_targets=uses_individual_targets,
        uses_emergency_range=uses_emergency_range,
        capacity_basis_label=capacity_basis_label,
        message=message,
    )


def _validate_pre_solver_conditions(req: ScheduleRequest, days_in_month: int) -> None:
    _validate_target_hours_within_work_period(req, days_in_month)
    _validate_consecutive_day_work_limits(req, days_in_month)
    _validate_daily_staffing_capacity(req, days_in_month)


def _validate_target_hours_within_work_period(req: ScheduleRequest, days_in_month: int) -> None:
    for index, worker in enumerate(req.workers):
        target_hours = 160 if worker.target_hours is None else int(worker.target_hours)
        start_day = max(1, int(worker.start_day or 1))
        end_day = int(worker.end_day if worker.end_day is not None else days_in_month)
        active_start = max(1, start_day)
        active_end = min(days_in_month, end_day)
        active_days = max(0, active_end - active_start + 1)
        min_hours = _worker_min_fixed_credit_hours(worker, days_in_month, req.settings.special_shifts)
        assignable_credit_slots = _worker_assignable_credit_slots(worker, days_in_month)
        max_hours = _worker_max_credit_hours(worker, days_in_month, req.settings.special_shifts)

        if min_hours > target_hours:
            fixed_summary = _worker_fixed_credit_summary(worker, days_in_month, req.settings.special_shifts)
            label = display_worker_name(worker.name, index)
            raise SchedulerError(
                f"{label}: 이미 고정된 연가/근무 인정시간이 {min_hours}시간으로 "
                f"목표시간 {target_hours}시간보다 많습니다. "
                f"현재 고정 입력은 {fixed_summary}입니다. "
                "연가/기타 근무/고정 근무를 줄이거나 목표시간을 늘려 주세요."
            )

        if target_hours <= max_hours:
            continue

        remaining_hours = max(0, target_hours - min_hours)
        required_slots = math.ceil(remaining_hours / 8) if remaining_hours > 0 else 0
        label = display_worker_name(worker.name, index)
        raise SchedulerError(
            f"{label}: 근무 가능 기간은 {start_day}~{end_day}일, 총 {active_days}일입니다. "
            f"목표시간 {target_hours}시간은 근무 시간이 초과됩니다. "
            f"고정 인정시간을 제외하고 {remaining_hours}시간을 더 채워야 하지만 "
            f"주/야/비로 채울 수 있는 칸은 {assignable_credit_slots}개뿐입니다. "
            f"8시간 기준 {required_slots}개 칸이 필요하므로 현재 최대 인정 가능 시간은 {max_hours}시간입니다. "
            "시작일/종료일을 넓히거나 휴무를 줄이거나 목표시간, 연가, 기타 근무 입력을 조정해 주세요."
        )


def _worker_min_fixed_credit_hours(
    worker: WorkerInput,
    days_in_month: int,
    special_shifts: dict[str, int],
) -> int:
    start_day = max(1, int(worker.start_day or 1))
    end_day = int(worker.end_day if worker.end_day is not None else days_in_month)
    if end_day < start_day:
        return 0

    total = 0
    for day in range(max(1, start_day), min(days_in_month, end_day) + 1):
        shift = _fixed_shift_for_day(worker, day)
        if shift in ("day", "night", "off_night", "leave"):
            total += 8
        elif shift == "off" or not shift:
            continue
        else:
            token = str(shift).strip()
            if token not in special_shifts:
                raise SchedulerError(f"기타 근무 '{token}'의 인정 시간이 설정에 없습니다.")
            total += int(special_shifts[token])
    return total


def _worker_fixed_credit_summary(
    worker: WorkerInput,
    days_in_month: int,
    special_shifts: dict[str, int],
) -> str:
    counts = {
        "주간": 0,
        "야간": 0,
        "비번": 0,
        "연가": 0,
        "기타 근무": 0,
    }
    for day in range(1, days_in_month + 1):
        shift = _fixed_shift_for_day(worker, day)
        if shift == "day":
            counts["주간"] += 1
        elif shift == "night":
            counts["야간"] += 1
        elif shift == "off_night":
            counts["비번"] += 1
        elif shift == "leave":
            counts["연가"] += 1
        elif shift and shift != "off":
            token = str(shift).strip()
            if token not in special_shifts:
                raise SchedulerError(f"기타 근무 '{token}'의 인정 시간이 설정에 없습니다.")
            counts["기타 근무"] += 1

    parts = [f"{label} {count}일" for label, count in counts.items() if count]
    return ", ".join(parts) if parts else "없음"


def _worker_assignable_credit_slots(worker: WorkerInput, days_in_month: int) -> int:
    count = 0
    start_day = max(1, int(worker.start_day or 1))
    end_day = int(worker.end_day if worker.end_day is not None else days_in_month)
    for day in range(max(1, start_day), min(days_in_month, end_day) + 1):
        if not _fixed_shift_for_day(worker, day):
            count += 1
    return count


def _worker_max_credit_hours(
    worker: WorkerInput,
    days_in_month: int,
    special_shifts: dict[str, int],
) -> int:
    start_day = max(1, int(worker.start_day or 1))
    end_day = int(worker.end_day if worker.end_day is not None else days_in_month)
    if end_day < start_day:
        return 0

    total = 0
    for day in range(max(1, start_day), min(days_in_month, end_day) + 1):
        shift = _fixed_shift_for_day(worker, day)
        if shift == "off":
            continue
        if shift and shift not in FIXED_SHIFT_CODES:
            token = str(shift).strip()
            if token not in special_shifts:
                raise SchedulerError(f"기타 근무 '{token}'의 인정 시간이 설정에 없습니다.")
            total += int(special_shifts[token])
        else:
            total += 8
    return total


def _validate_consecutive_day_work_limits(req: ScheduleRequest, days_in_month: int) -> None:
    max_allowed = 5
    for index, worker in enumerate(req.workers):
        day_work_days = _implied_day_work_days(worker, days_in_month, req.settings)
        longest = _longest_consecutive_run(day_work_days)
        if longest > max_allowed:
            label = _rule_worker_label(worker, index)
            raise SchedulerError(f"{label}: 주간 연속 일수가 {longest}일로 규칙에 위배됩니다.")


def _implied_day_work_days(
    worker: WorkerInput,
    days_in_month: int,
    settings: ScheduleSettings,
) -> set[int]:
    day_work_days = {
        day
        for day in range(1, days_in_month + 1)
        if _fixed_shift_for_day(worker, day) == "day"
    }

    if _dedicated_shift_from_input(worker) == "night" or _effective_max_day_staffing(settings) <= 0:
        return day_work_days

    target_hours = 160 if worker.target_hours is None else int(worker.target_hours)
    fixed_credit_hours = _worker_min_fixed_credit_hours(worker, days_in_month, settings.special_shifts)
    remaining_hours = target_hours - fixed_credit_hours
    if remaining_hours <= 0 or remaining_hours % 8 != 0:
        return day_work_days

    assignable_days = _worker_assignable_credit_days(worker, days_in_month)
    required_regular_days = remaining_hours // 8
    if required_regular_days != len(assignable_days):
        return day_work_days

    for run_start, run_end in _consecutive_runs(assignable_days):
        tail_non_day_credit_days = _credit_non_day_tail_capacity(worker, run_start, run_end, days_in_month, settings)
        implied_day_end = run_end - tail_non_day_credit_days
        if implied_day_end >= run_start:
            day_work_days.update(range(run_start, implied_day_end + 1))

    return day_work_days


def _effective_max_day_staffing(settings: ScheduleSettings) -> int:
    max_day = int(settings.max_day)
    if settings.use_emergency_range and settings.emergency_max_day is not None:
        max_day = max(max_day, int(settings.emergency_max_day))
    return max_day


def _worker_assignable_credit_days(worker: WorkerInput, days_in_month: int) -> list[int]:
    start_day = max(1, int(worker.start_day or 1))
    end_day = int(worker.end_day if worker.end_day is not None else days_in_month)
    return [
        day
        for day in range(max(1, start_day), min(days_in_month, end_day) + 1)
        if not _fixed_shift_for_day(worker, day)
    ]


def _consecutive_runs(days: Iterable[int]) -> list[tuple[int, int]]:
    runs: list[tuple[int, int]] = []
    start: int | None = None
    previous: int | None = None
    for day in sorted(days):
        if start is None:
            start = day
            previous = day
            continue
        if previous is not None and day == previous + 1:
            previous = day
            continue
        runs.append((start, previous if previous is not None else start))
        start = day
        previous = day
    if start is not None:
        runs.append((start, previous if previous is not None else start))
    return runs


def _credit_non_day_tail_capacity(
    worker: WorkerInput,
    run_start: int,
    run_end: int,
    days_in_month: int,
    settings: ScheduleSettings,
) -> int:
    run_length = run_end - run_start + 1
    if run_length < 2 or _dedicated_shift_from_input(worker) == "day":
        return 0

    if (
        settings.allow_double_night_cycle
        and run_length >= 4
        and _can_end_run_with_double_night_cycle(worker, run_end, days_in_month, settings)
    ):
        return 4

    if _can_end_run_with_off_night(worker, run_end, days_in_month, settings):
        return 2
    return 0


def _can_end_run_with_double_night_cycle(
    worker: WorkerInput,
    run_end: int,
    days_in_month: int,
    settings: ScheduleSettings,
) -> bool:
    first_off_night_day = run_end - 2
    if first_off_night_day < 2 or run_end + 2 > days_in_month:
        return False
    return (
        _is_allowed_after_off_night_fixed_shift(_fixed_shift_for_day(worker, run_end + 1), settings)
        and _is_allowed_after_off_night_fixed_shift(_fixed_shift_for_day(worker, run_end + 2), settings)
    )


def _can_end_run_with_off_night(
    worker: WorkerInput,
    run_end: int,
    days_in_month: int,
    settings: ScheduleSettings,
) -> bool:
    if run_end >= days_in_month:
        return True
    return _is_allowed_after_off_night_fixed_shift(_fixed_shift_for_day(worker, run_end + 1), settings)


def _is_allowed_after_off_night_fixed_shift(shift: str, settings: ScheduleSettings) -> bool:
    if shift == "off":
        return True
    return bool(settings.allow_leave_after_off_night and shift == "leave")


def _longest_consecutive_run(days: set[int]) -> int:
    longest = 0
    for start, end in _consecutive_runs(days):
        longest = max(longest, end - start + 1)
    return longest


def _rule_worker_label(worker: WorkerInput, index: int) -> str:
    name = str(worker.name or "").strip()
    if name:
        return name
    return display_worker_name(worker.name, index)


def _validate_daily_staffing_capacity(req: ScheduleRequest, days_in_month: int) -> None:
    min_day, min_night, basis_label = _effective_min_staffing(req.settings)
    if min_day <= 0 and min_night <= 0:
        return

    reasons: list[str] = []
    for day in range(1, days_in_month + 1):
        stats = _daily_staffing_capacity(req, day, days_in_month)
        fixed_day = stats["fixed_day"]
        fixed_night = stats["fixed_night"]
        fixed_off_night = stats["fixed_off_night"]
        required_off_night = _required_off_night_for_day(req, day, days_in_month, min_night)
        day_need = max(0, min_day - fixed_day)
        night_need = max(0, min_night - fixed_night)
        off_night_need = max(0, required_off_night - fixed_off_night)
        day_only = stats["day_only"]
        night_only = stats["night_only"]
        both = stats["both"]
        day_possible = fixed_day + day_only + both
        night_possible = fixed_night + night_only + both
        off_night_possible = fixed_off_night + night_only + both
        combined_possible = fixed_day + fixed_night + fixed_off_night + day_only + night_only + both
        combined_needed = min_day + min_night + required_off_night

        day_reasons: list[str] = []
        if day_possible < min_day:
            day_reasons.append(f"주간 최소 {min_day}명 중 최대 {day_possible}명만 가능")
        if night_possible < min_night:
            day_reasons.append(f"야간 최소 {min_night}명 중 최대 {night_possible}명만 가능")
        if off_night_possible < required_off_night:
            day_reasons.append(f"전날 야간에 따른 비번 최소 {required_off_night}명 중 최대 {off_night_possible}명만 가능")
        if night_need + off_night_need > night_only + both:
            day_reasons.append(
                f"야간/비번 필요 {night_need + off_night_need}명 중 최대 {night_only + both}명만 가능"
            )
        if day_need + night_need + off_night_need > day_only + night_only + both:
            day_reasons.append(
                f"남은 주/야/비 필요 {day_need + night_need + off_night_need}명 중 최대 {day_only + night_only + both}명만 가능"
            )
        if combined_possible < combined_needed:
            day_reasons.append(f"주간+야간+비번 최소 {combined_needed}명 중 최대 {combined_possible}명만 가능")

        if not day_reasons:
            continue

        unavailable_text = _daily_unavailable_summary(stats)
        reasons.append(
            f"{day}일: {basis_label} 기준으로 {', '.join(day_reasons)}합니다. "
            f"{unavailable_text} 휴무/연가/비번/기타 근무 입력이나 기본 TO 설정을 조정해 주세요."
        )
        if len(reasons) >= 3:
            break

    if reasons:
        raise SchedulerError("날짜별 최소 인원을 채울 수 없습니다.\n" + "\n".join(reasons))


def _effective_min_staffing(settings: ScheduleSettings) -> tuple[int, int, str]:
    if settings.use_emergency_range:
        min_day = int(settings.emergency_min_day if settings.emergency_min_day is not None else settings.min_day)
        min_night = int(settings.emergency_min_night if settings.emergency_min_night is not None else settings.min_night)
        return min_day, min_night, "예외 범위"
    return int(settings.min_day), int(settings.min_night), "기본 범위"


def _daily_staffing_capacity(req: ScheduleRequest, day: int, days_in_month: int) -> dict[str, object]:
    stats: dict[str, object] = {
        "fixed_day": 0,
        "fixed_night": 0,
        "fixed_off_night": 0,
        "day_only": 0,
        "night_only": 0,
        "both": 0,
        "unavailable": [],
    }

    for index, worker in enumerate(req.workers):
        label = display_worker_name(worker.name, index)
        start_day = max(1, int(worker.start_day or 1))
        end_day = int(worker.end_day if worker.end_day is not None else days_in_month)
        if day < start_day or day > end_day:
            stats["unavailable"].append(f"{label}(기간 외)")
            continue

        shift = _fixed_shift_for_day(worker, day)
        previous_shift = _fixed_shift_for_day(worker, day - 1) if day > 1 else ""
        forced_off_night = (day == 1 and worker.prev_month_last_day_night) or previous_shift == "night"

        if shift == "day":
            stats["fixed_day"] += 1
            continue
        if shift == "night":
            stats["fixed_night"] += 1
            continue
        if shift == "off_night" or forced_off_night:
            stats["fixed_off_night"] += 1
            stats["unavailable"].append(f"{label}(비번)")
            continue
        if shift == "off":
            stats["unavailable"].append(f"{label}(휴무)")
            continue
        if shift == "leave":
            stats["unavailable"].append(f"{label}(연가)")
            continue
        if shift:
            stats["unavailable"].append(f"{label}(기타 근무)")
            continue

        dedicated_shift = _dedicated_shift_from_input(worker)
        if dedicated_shift == "day":
            stats["day_only"] += 1
        elif dedicated_shift == "night":
            stats["night_only"] += 1
        else:
            stats["both"] += 1

    return stats


def _required_off_night_for_day(
    req: ScheduleRequest,
    day: int,
    days_in_month: int,
    min_night: int,
) -> int:
    if day <= 1:
        return sum(1 for worker in req.workers if worker.prev_month_last_day_night)

    previous_day = day - 1
    fixed_previous_night = sum(
        1
        for worker in req.workers
        if _fixed_shift_for_day(worker, previous_day) == "night"
        and _is_worker_active_input_on_day(worker, previous_day, days_in_month)
    )
    return max(0, min_night, fixed_previous_night)


def _is_worker_active_input_on_day(worker: WorkerInput, day: int, days_in_month: int) -> bool:
    start_day = max(1, int(worker.start_day or 1))
    end_day = int(worker.end_day if worker.end_day is not None else days_in_month)
    return start_day <= day <= end_day


def _daily_unavailable_summary(stats: dict[str, object]) -> str:
    unavailable = list(stats.get("unavailable", []))
    if not unavailable:
        return "배치 제외 인원은 없습니다."
    shown = ", ".join(str(item) for item in unavailable[:5])
    suffix = f" 외 {len(unavailable) - 5}명" if len(unavailable) > 5 else ""
    return f"배치 제외 인원: {shown}{suffix}."


def _fixed_shift_for_day(worker: WorkerInput, day: int) -> str:
    fixed = worker.fixed_shifts or {}
    return str(fixed.get(day, fixed.get(str(day), "")) or "").strip()


def _total_target_hours(worker_inputs: Iterable[WorkerInput]) -> int:
    total = 0
    for worker in worker_inputs:
        total += 160 if worker.target_hours is None else int(worker.target_hours)
    return total


def _non_to_credit_hours(req: ScheduleRequest) -> tuple[int, int]:
    leave_credit_hours = 0
    special_credit_hours = 0
    special_shifts = req.settings.special_shifts

    for worker in req.workers:
        for day_number, shift in worker.fixed_shifts.items():
            int(day_number)  # Validate that the key is numeric enough for downstream use.
            if shift == "leave":
                leave_credit_hours += 8
            elif shift in FIXED_SHIFT_CODES:
                continue
            else:
                token = str(shift).strip()
                if not token:
                    continue
                if token not in special_shifts:
                    raise SchedulerError(f"기타 근무 '{token}'의 인정 시간이 설정에 없습니다.")
                special_credit_hours += int(special_shifts[token])

    return leave_credit_hours, special_credit_hours


def _max_regular_capacity_hours(settings, days_in_month: int) -> int:
    max_day = int(settings.max_day)
    max_night = int(settings.max_night)
    if settings.use_emergency_range:
        if settings.emergency_max_day is not None:
            max_day = max(max_day, int(settings.emergency_max_day))
        if settings.emergency_max_night is not None:
            max_night = max(max_night, int(settings.emergency_max_night))
    day_capacity = max_day * days_in_month * 8
    night_capacity = max_night * days_in_month * 8
    generated_off_night_capacity = max_night * max(0, days_in_month - 1) * 8
    return day_capacity + night_capacity + generated_off_night_capacity


def export_schedule_result_to_excel(
    template_path: str,
    output_path: str,
    result: ScheduleResult,
    apply_shift_colors: bool = False,
) -> ExcelExportResult:
    workers = [Worker(id=idx, name=row.name) for idx, row in enumerate(result.rows)]
    schedule = [list(row.raw_days) for row in result.rows]

    export_to_excel(
        template_path=template_path,
        output_path=output_path,
        workers=workers,
        schedule=schedule,
        num_days=result.days_in_month,
        year=result.year,
        month=result.month,
        apply_shift_colors=False,
    )

    return ExcelExportResult(
        output_path=output_path,
        filename=Path(output_path).name or default_excel_filename(result.year, result.month),
        worker_count=len(workers),
        days_in_month=result.days_in_month,
    )


def default_excel_filename(year: int, month: int) -> str:
    return f"근무표_{year}_{month}.xlsx"


def display_worker_name(name: str, index: int) -> str:
    stripped = str(name or "").strip()
    if stripped:
        return stripped
    return excel_style_label(index)


def excel_style_label(index: int) -> str:
    if index < 0:
        raise ValueError("index must be non-negative")

    value = index + 1
    chars: list[str] = []
    while value:
        value, remainder = divmod(value - 1, 26)
        chars.append(chr(ord("A") + remainder))
    return "".join(reversed(chars))
